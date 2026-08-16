"""
Load order information from Excel.
"""
from __future__ import annotations

import re
import time

from openpyxl import load_workbook

from models.order_model import Order
from project_paths import ensure_managed_data_file, resolve_project_path


PRODUCT_HEADER_RE = re.compile(r"^\[상품(\d+)\]")
RECEIPT_HEADER = "수령확인"
SEAT_HEADER = "좌석번호"
ORDER_STATUS_HEADER = "주문상태"
_WRITE_RETRY_COUNT = 3
_WRITE_RETRY_DELAY_SEC = 0.2


class ExcelService:
    def __init__(self, file_path: str | None = None):
        if file_path is None:
            self._file_path = str(ensure_managed_data_file())
        else:
            self._file_path = str(resolve_project_path(file_path))

    def search_orders(self, keyword: str = "") -> list[Order]:
        """주문번호/이름/연락처로 부분 일치 검색. 빈 키워드면 전체 반환(최대 200건)."""
        keyword = keyword.strip()
        results: list[Order] = []
        max_results = 200

        workbook = load_workbook(self._file_path, read_only=True, data_only=True)
        try:
            ws = workbook.active
            headers = self._read_headers(ws)

            order_col = self._find_col(headers, ("주문번호",))
            if not order_col:
                return results

            name_col = self._find_col(headers, ("주문자명", "수령자명"))
            phone_col = self._find_col(headers, ("주문자연락처", "수령자연락처"))
            seat_col = self._find_col(headers, ("좌석번호",))
            received_col = self._find_col(headers, (RECEIPT_HEADER,))
            status_col = self._find_col(headers, (ORDER_STATUS_HEADER,))
            goods_cols = self._parse_goods_cols(headers)

            for row in ws.iter_rows(min_row=2, values_only=True):
                order_number_val = str(self._cell(row, order_col)).strip()
                if not order_number_val:
                    continue

                name_val = str(self._cell(row, name_col)).strip() if name_col else ""
                phone_val = str(self._cell(row, phone_col)).strip() if phone_col else ""

                if keyword and not any(
                    keyword in field for field in (order_number_val, name_val, phone_val)
                ):
                    continue

                goods_list = self._build_goods_list(row, goods_cols)
                results.append(Order(
                    order_number=order_number_val,
                    name=name_val,
                    phone=phone_val,
                    seat=str(self._cell(row, seat_col)).strip() if seat_col else "",
                    goods=goods_list,
                    received_at=str(self._cell(row, received_col)).strip() if received_col else "",
                    order_status=str(self._cell(row, status_col)).strip() if status_col else "",
                ))
                if len(results) >= max_results:
                    break

            return results
        finally:
            workbook.close()

    def find_order(self, order_number: str) -> Order | None:
        """Find order by order_number."""
        workbook = load_workbook(self._file_path, read_only=True, data_only=True)
        try:
            ws = workbook.active
            headers = self._read_headers(ws)

            order_col = self._find_col(headers, ("주문번호",))
            if not order_col:
                return None

            name_col = self._find_col(headers, ("주문자명", "수령자명"))
            phone_col = self._find_col(headers, ("주문자연락처", "수령자연락처"))
            seat_col = self._find_col(headers, ("좌석번호",))
            received_col = self._find_col(headers, (RECEIPT_HEADER,))
            status_col = self._find_col(headers, (ORDER_STATUS_HEADER,))

            goods_cols = self._parse_goods_cols(headers)

            for row in ws.iter_rows(min_row=2, values_only=True):
                current_order = self._cell(row, order_col)
                if str(current_order).strip().upper() != order_number.upper():
                    continue

                return Order(
                    order_number=order_number,
                    name=str(self._cell(row, name_col)).strip() if name_col else "",
                    phone=str(self._cell(row, phone_col)).strip() if phone_col else "",
                    seat=str(self._cell(row, seat_col)).strip() if seat_col else "",
                    goods=self._build_goods_list(row, goods_cols),
                    received_at=str(self._cell(row, received_col)).strip() if received_col else "",
                    order_status=str(self._cell(row, status_col)).strip() if status_col else "",
                )

            return None
        finally:
            workbook.close()

    def find_orders_by_customer(self, name: str = "", phone: str = "") -> list[Order]:
        """Find orders by exact customer name/phone match."""
        normalized_name = (name or "").strip()
        normalized_phone = self._normalize_phone(phone)
        if not normalized_name and not normalized_phone:
            return []

        workbook = load_workbook(self._file_path, read_only=True, data_only=True)
        try:
            ws = workbook.active
            headers = self._read_headers(ws)

            order_col = self._find_col(headers, ("주문번호",))
            if not order_col:
                return []

            name_col = self._find_col(headers, ("주문자명",))
            phone_col = self._find_col(headers, ("주문자연락처",))
            recv_name_col = self._find_col(headers, ("수령자명",))
            recv_phone_col = self._find_col(headers, ("수령자연락처",))
            seat_col = self._find_col(headers, ("좌석번호",))
            received_col = self._find_col(headers, (RECEIPT_HEADER,))
            goods_cols = self._parse_goods_cols(headers)

            matches: list[Order] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                order_number = str(self._cell(row, order_col)).strip()
                if not order_number:
                    continue

                names = {
                    str(self._cell(row, name_col)).strip() if name_col else "",
                    str(self._cell(row, recv_name_col)).strip() if recv_name_col else "",
                }
                phones = {
                    self._normalize_phone(self._cell(row, phone_col)) if phone_col else "",
                    self._normalize_phone(self._cell(row, recv_phone_col)) if recv_phone_col else "",
                }

                name_matches = not normalized_name or normalized_name in names
                phone_matches = not normalized_phone or normalized_phone in phones
                if not name_matches or not phone_matches:
                    continue

                matches.append(Order(
                    order_number=order_number,
                    name=str(self._cell(row, name_col)).strip() if name_col else "",
                    phone=str(self._cell(row, phone_col)).strip() if phone_col else "",
                    seat=str(self._cell(row, seat_col)).strip() if seat_col else "",
                    goods=self._build_goods_list(row, goods_cols),
                    received_at=str(self._cell(row, received_col)).strip() if received_col else "",
                ))

            return matches
        finally:
            workbook.close()

    def find_unique_order_by_customer(self, name: str = "", phone: str = "") -> Order | None:
        matches = self.find_orders_by_customer(name=name, phone=phone)
        if len(matches) != 1:
            return None
        return matches[0]

    def _ensure_column(self, header_name: str) -> None:
        """지정한 헤더 컬럼이 없으면 자동 추가한다."""
        for attempt in range(_WRITE_RETRY_COUNT):
            workbook = None
            try:
                workbook = load_workbook(self._file_path)
                ws = workbook.active
                headers = self._read_headers(ws)
                if self._find_col(headers, (header_name,)):
                    return
                new_col = ws.max_column + 1
                ws.cell(row=1, column=new_col, value=header_name)
                workbook.save(self._file_path)
                return
            except (PermissionError, OSError):
                if attempt == _WRITE_RETRY_COUNT - 1:
                    return
                time.sleep(_WRITE_RETRY_DELAY_SEC)
            finally:
                if workbook is not None:
                    workbook.close()

    def ensure_seat_column(self) -> None:
        """data.xlsx에 좌석번호 컬럼이 없으면 자동 추가한다."""
        self._ensure_column(SEAT_HEADER)

    def ensure_receipt_column(self) -> None:
        """data.xlsx에 수령확인 컬럼이 없으면 자동 추가한다."""
        self._ensure_column(RECEIPT_HEADER)

    def ensure_order_status_column(self) -> None:
        """data.xlsx에 주문상태 컬럼이 없으면 자동 추가한다."""
        self._ensure_column(ORDER_STATUS_HEADER)

    def mark_order_status(self, order_number: str, status: str) -> bool:
        """주문의 주문상태 값을 엑셀에 저장한다."""
        for attempt in range(_WRITE_RETRY_COUNT):
            workbook = None
            try:
                workbook = load_workbook(self._file_path)
                ws = workbook.active
                headers = self._read_headers(ws)
                order_col = self._find_col(headers, ("주문번호",))
                if not order_col:
                    return False
                status_col = self._find_col(headers, (ORDER_STATUS_HEADER,))
                if not status_col:
                    status_col = ws.max_column + 1
                    ws.cell(row=1, column=status_col, value=ORDER_STATUS_HEADER)
                target_row = self._find_row_by_order(ws, order_col, order_number)
                if not target_row:
                    return False
                ws.cell(row=target_row, column=status_col, value=(status or "").strip())
                workbook.save(self._file_path)
                return True
            except (PermissionError, OSError):
                if attempt == _WRITE_RETRY_COUNT - 1:
                    return False
                time.sleep(_WRITE_RETRY_DELAY_SEC)
            finally:
                if workbook is not None:
                    workbook.close()
        return False

    def get_received_status_map(self) -> dict[str, str]:
        """현재 파일에서 수령확인이 기록된 주문번호 → 타임스탬프 맵을 반환한다."""
        workbook = None
        try:
            workbook = load_workbook(self._file_path, read_only=True, data_only=True)
            ws = workbook.active
            headers = self._read_headers(ws)
            order_col = self._find_col(headers, ("주문번호",))
            received_col = self._find_col(headers, (RECEIPT_HEADER,))
            if not order_col or not received_col:
                return {}
            result: dict[str, str] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                order_number = str(self._cell(row, order_col)).strip()
                received = str(self._cell(row, received_col)).strip()
                if order_number and received:
                    result[order_number] = received
            return result
        except Exception:
            return {}
        finally:
            if workbook is not None:
                workbook.close()

    def bulk_restore_received_status(self, received_map: dict[str, str]) -> int:
        """파일 교체 후 수령확인 상태를 일괄 복원한다. 반환값: 복원된 건수."""
        if not received_map:
            return 0
        for attempt in range(_WRITE_RETRY_COUNT):
            workbook = None
            try:
                workbook = load_workbook(self._file_path)
                ws = workbook.active
                headers = self._read_headers(ws)
                order_col = self._find_col(headers, ("주문번호",))
                if not order_col:
                    return 0
                receipt_col = self._find_col(headers, (RECEIPT_HEADER,))
                if not receipt_col:
                    receipt_col = ws.max_column + 1
                    ws.cell(row=1, column=receipt_col, value=RECEIPT_HEADER)
                count = 0
                for row_idx in range(2, ws.max_row + 1):
                    order_number = str(ws.cell(row=row_idx, column=order_col).value or "").strip()
                    if order_number in received_map:
                        ws.cell(row=row_idx, column=receipt_col, value=received_map[order_number])
                        count += 1
                workbook.save(self._file_path)
                return count
            except (PermissionError, OSError):
                if attempt == _WRITE_RETRY_COUNT - 1:
                    return 0
                time.sleep(_WRITE_RETRY_DELAY_SEC)
            finally:
                if workbook is not None:
                    workbook.close()
        return 0

    def get_product_names(self) -> list[str]:
        """상품 컬럼명 리스트를 반환한다 (티켓 분류 UI용)."""
        workbook = load_workbook(self._file_path, read_only=True, data_only=True)
        try:
            ws = workbook.active
            headers = self._read_headers(ws)
            goods_cols = self._parse_goods_cols(headers)
            return [name or f"상품{idx}" for idx, _col, name in goods_cols]
        finally:
            workbook.close()

    def mark_order_received(self, order_number: str, timestamp_str: str) -> bool:
        """Mark order as received and persist timestamp."""
        return self._update_order_received(order_number, timestamp_str)

    def rollback_order_received(self, order_number: str, previous_value: str) -> bool:
        """Restore previous receipt value when print step fails."""
        return self._update_order_received(order_number, previous_value)

    def _update_order_received(self, order_number: str, value: str) -> bool:
        for attempt in range(_WRITE_RETRY_COUNT):
            workbook = None
            try:
                workbook = load_workbook(self._file_path)
                ws = workbook.active

                headers = self._read_headers(ws)
                order_col = self._find_col(headers, ("주문번호",))
                if not order_col:
                    return False

                receipt_col = self._find_col(headers, (RECEIPT_HEADER,))
                if not receipt_col:
                    receipt_col = ws.max_column + 1
                    ws.cell(row=1, column=receipt_col, value=RECEIPT_HEADER)

                target_row = self._find_row_by_order(ws, order_col, order_number)
                if not target_row:
                    return False

                ws.cell(row=target_row, column=receipt_col, value=(value or "").strip())
                workbook.save(self._file_path)
                return True
            except (PermissionError, OSError):
                if attempt == _WRITE_RETRY_COUNT - 1:
                    return False
                time.sleep(_WRITE_RETRY_DELAY_SEC)
            finally:
                if workbook is not None:
                    workbook.close()
        return False

    @staticmethod
    def _parse_goods_cols(headers: dict[str, int]) -> list[tuple[int, int, str]]:
        """상품 헤더를 파싱해 (상품번호, 컬럼인덱스, 상품명) 리스트를 반환한다."""
        goods_cols: list[tuple[int, int, str]] = []
        for header_name, col_idx in headers.items():
            match = PRODUCT_HEADER_RE.match(header_name)
            if not match:
                continue
            product_index = int(match.group(1))
            clean_name = PRODUCT_HEADER_RE.sub("", header_name).strip()
            goods_cols.append((product_index, col_idx, clean_name))
        goods_cols.sort(key=lambda item: item[0])
        return goods_cols

    @staticmethod
    def _build_goods_list(row: tuple, goods_cols: list[tuple[int, int, str]]) -> list[str]:
        """행에서 상품 목록 문자열 리스트를 생성한다."""
        goods_list: list[str] = []
        for product_index, col_idx, goods_name in goods_cols:
            quantity = ExcelService._to_int(ExcelService._cell(row, col_idx))
            if quantity > 0:
                label = goods_name or f"상품{product_index}"
                goods_list.append(f"{label} x{quantity}")
        return goods_list

    @staticmethod
    def _read_headers(ws) -> dict[str, int]:
        header_row = [cell.value for cell in ws[1]]
        return {
            str(value).strip(): idx
            for idx, value in enumerate(header_row, 1)
            if value is not None and str(value).strip()
        }

    @staticmethod
    def _find_row_by_order(ws, order_col: int, order_number: str) -> int | None:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            current_order = ExcelService._cell(row, order_col)
            if str(current_order).strip().upper() == order_number.upper():
                return row_idx
        return None

    @staticmethod
    def _find_col(headers: dict[str, int], candidates: tuple[str, ...]) -> int | None:
        for candidate in candidates:
            if candidate in headers:
                return headers[candidate]

        normalized = {key.replace(" ", ""): idx for key, idx in headers.items()}
        for candidate in candidates:
            key = candidate.replace(" ", "")
            if key in normalized:
                return normalized[key]
        return None

    @staticmethod
    def _cell(row: tuple, col_idx: int | None):
        if not col_idx:
            return ""
        idx = col_idx - 1
        if idx < 0 or idx >= len(row):
            return ""
        value = row[idx]
        return "" if value is None else value

    @staticmethod
    def _to_int(value) -> int:
        if value is None or value == "":
            return 0
        try:
            return int(float(str(value).strip()))
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _normalize_phone(value) -> str:
        raw = str(value).strip() if value is not None else ""
        return re.sub(r"\D+", "", raw)
