"""두 Excel 파일의 앞부분을 JSON으로 비교하는 일회성 진단 도구."""

import json

import openpyxl


wb1 = openpyxl.load_workbook("data.xlsx", data_only=True)
ws1 = wb1.active
data = {"data_rows": []}
for row in ws1.iter_rows(min_row=1, max_row=3, values_only=True):
    data["data_rows"].append([str(c) if c is not None else "" for c in row])

wb2 = openpyxl.load_workbook("Witchform_Payform_20260306_183537.xlsx", data_only=True)
ws2 = wb2.active
data["witch_rows"] = []
for row in ws2.iter_rows(min_row=1, max_row=3, values_only=True):
    data["witch_rows"].append([str(c) if c is not None else "" for c in row])

with open("output_rows.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
