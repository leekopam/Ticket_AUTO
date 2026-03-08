from pyzbar.pyzbar import decode
import cv2
import requests
from urllib.parse import urlparse, parse_qs
from openpyxl import load_workbook
import tkinter as tk
from tkinter import font
import threading
import webbrowser
import os

processing = False

def load_session():
    if os.path.exists('session.txt'):
        with open('session.txt', 'r') as f:
            return f.read().strip()
    return ""

PHPSESSID = load_session()

def show_error(message):
    global processing
    root = tk.Tk()
    root.title("오류")
    root.geometry("600x400")
    
    large_font = font.Font(size=20, weight="bold")
    
    tk.Label(root, text=message, font=large_font, fg="red").pack(pady=50)
    
    def on_close():
        global processing
        processing = False
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()

def show_order_info(order_number, url):
    global processing
    wb = load_workbook('data.xlsx')
    ws = wb.active
    
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    
    if "주문번호" not in headers:
        show_error("주문번호 열을 찾을 수 없습니다")
        return
    
    order_col = headers["주문번호"]
    name_col = headers.get("주문자명")
    phone_col = headers.get("주문자연락처")
    seat_col = headers.get("좌석번호")
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[order_col-1].value == order_number:
            name = row[name_col-1].value if name_col else ""
            phone = row[phone_col-1].value if phone_col else ""
            seat = row[seat_col-1].value if seat_col else ""
            
            root = tk.Tk()
            root.title("주문 정보")
            root.geometry("600x400")
            
            large_font = font.Font(size=20, weight="bold")
            
            tk.Label(root, text=f"주문자명: {name}", font=large_font).pack(pady=30)
            tk.Label(root, text=f"주문자연락처: {phone}", font=large_font).pack(pady=30)
            tk.Label(root, text=f"좌석번호: {seat}", font=large_font).pack(pady=30)
            
            webbrowser.open(url)
            
            def on_close():
                global processing
                processing = False
                root.destroy()
            
            root.protocol("WM_DELETE_WINDOW", on_close)
            root.mainloop()
            return
    
    show_error(f"주문번호를 찾을 수 없습니다")

def process_qr(url):
    global processing
    
    print(f"QR Code: {url}")
    
    #판매자가 맞는지 검증
    if url.startswith("https://witchform.com/qrcode_link.php"):
        processing = True
        
        cookies = {"PHPSESSID": PHPSESSID}
        response = requests.get(url, cookies=cookies, allow_redirects=False)
        
        if response.status_code in [301, 302, 303, 307, 308]:
            location = response.headers.get('Location')
            print(f"Redirect to: {location}")
            
            parsed = urlparse(location)
            path = parsed.path
            
            if path.startswith("/w/myform/sellForm-history-detail"):
                path_parts = path.split('/')
                last_part = path_parts[-1]
                
                query_params = parse_qs(parsed.query)
                uuid = query_params.get('idx', [''])[0]
                
                order_number = f"{uuid}_{last_part}"
                print(f"Order Number: {order_number}")
                
                full_url = f"https://witchform.com{location}" if location.startswith('/') else location
                # 멀티스레드로 show_order_info(order_number, full_url) 메서드 실행, 
                # 그냥 호출하면 웹캡 스캔하는게 멈춤
                threading.Thread(target=show_order_info, args=(order_number, full_url)).start()
            else:
                threading.Thread(target=show_error, args=("잘못된 경로",)).start()
        else:
            threading.Thread(target=show_error, args=("인증 오류",)).start()
    else:
        print(url)

def scan_webcam():
    cap = cv2.VideoCapture(0)
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        
        if not processing:
            qr_codes = decode(frame)
            
            for qr in qr_codes:
                url = qr.data.decode('utf-8')
                process_qr(url)
                break
        
        cv2.imshow('QR Scanner', frame)
        
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    
    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    scan_webcam() 
