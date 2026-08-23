"""Main receipt flow tests."""
from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from models.order_model import Order
from services.api_service import QRParseResult
from services.browser_service import BrowserResolveResult, ReceiptClickResult
from services.excel_service import ExcelService

import main as app_main


_USE_DEFAULT_LOAD_RESULT = object()


class _FakeScannerView:
    def __init__(self) -> None:
        self.auth_ready = True
        self.scanning_enabled = True
        self.status_message = ""
        self.started = False
        self.released = False

    def set_auth_ready(self, ready: bool) -> None:
        self.auth_ready = ready

    def set_scanning_enabled(self, enabled: bool) -> None:
        self.scanning_enabled = enabled

    def set_status_message(self, message: str) -> None:
        self.status_message = message

    def start(self) -> None:
        self.started = True

    def release(self) -> None:
        self.released = True

    def is_camera_ready(self) -> bool:
        return True

    def is_running(self) -> bool:
        return self.started and not self.released


class _FakeOrderView:
    def __init__(self) -> None:
        self.last_order: Order | None = None

    def show_or_update(self, order: Order) -> None:
        self.last_order = order


class _FakeOrderViewModel:
    def __init__(
        self,
        *,
        order: Order,
        click_result: ReceiptClickResult | None = None,
        click_result_seq: list[ReceiptClickResult] | None = None,
        load_result: Order | None | object = _USE_DEFAULT_LOAD_RESULT,
        open_ok: bool = True,
        open_exception: Exception | None = None,
        mark_ok: bool = True,
        rollback_ok: bool = True,
    ) -> None:
        self.order = order
        self.click_result = click_result or ReceiptClickResult(success=True)
        self._click_result_seq: list[ReceiptClickResult] = list(click_result_seq or [])
        self.load_result = order if load_result is _USE_DEFAULT_LOAD_RESULT else load_result
        self.open_ok = open_ok
        self.open_exception = open_exception
        self.mark_ok = mark_ok
        self.rollback_ok = rollback_ok
        self.load_calls = 0
        self.open_calls = 0
        self.complete_calls = 0
        self.mark_calls = 0
        self.rollback_calls = 0

    def load_order(self, order_number: str, url: str, *, open_page: bool = True) -> Order | None:
        self.load_calls += 1
        if self.load_result is None:
            return None
        self.order.order_number = order_number
        self.order.url = url
        return self.order

    def open_current_order_page(self) -> bool:
        self.open_calls += 1
        if self.open_exception is not None:
            raise self.open_exception
        return self.open_ok

    def complete_receipt(self) -> ReceiptClickResult:
        self.complete_calls += 1
        if self._click_result_seq:
            return self._click_result_seq.pop(0)
        return self.click_result

    def mark_current_order_received(self, timestamp_str: str | None = None) -> bool:
        self.mark_calls += 1
        if self.mark_ok:
            self.order.received_at = timestamp_str or ""
        return self.mark_ok

    def rollback_current_order_received(self, previous_value: str) -> bool:
        self.rollback_calls += 1
        if self.rollback_ok:
            self.order.received_at = previous_value
        return self.rollback_ok


class _FakeExcelService:
    def __init__(self, *, status_ok: bool = True, processing_time_ok: bool = True) -> None:
        self.status_ok = status_ok
        self.processing_time_ok = processing_time_ok
        self.status_calls: list[tuple[str, str]] = []
        self.processing_time_calls: list[tuple[str, str]] = []

    def mark_order_status(self, order_number: str, status: str) -> bool:
        self.status_calls.append((order_number, status))
        return self.status_ok

    def mark_order_processing_time(self, order_number: str, timestamp: str) -> bool:
        self.processing_time_calls.append((order_number, timestamp))
        return self.processing_time_ok


class _TrackingExcelService(ExcelService):
    def __init__(self, file_path: str) -> None:
        super().__init__(file_path)
        self.processing_time_calls: list[tuple[str, str]] = []

    def mark_order_processing_time(self, order_number: str, timestamp: str) -> bool:
        self.processing_time_calls.append((order_number, timestamp))
        return super().mark_order_processing_time(order_number, timestamp)


def _build_app_with_order_vm(order_vm: _FakeOrderViewModel):
    app = app_main.Application.__new__(app_main.Application)
    app._state = app_main.AppState.READY
    app._scanner_view = _FakeScannerView()
    app._order_view = _FakeOrderView()
    app._order_viewmodel = order_vm
    app._excel_service = _FakeExcelService()
    app._receipt_settings = SimpleNamespace(qr_scan_success_sound_path="")
    app._api_service = SimpleNamespace(
        parse_qr_redirect=lambda *_args, **_kwargs: QRParseResult(
            success=True,
            order_number="ORDER-001",
            full_url="https://example.com/order/1",
        )
    )
    app._order_listener = None
    app._status_listener = None
    app._stop_requested = False
    app._settings_store = SimpleNamespace(load=lambda: app._receipt_settings)
    app._audio_service = None
    app._ticket_debug_tools_service = SimpleNamespace(
        load_settings=lambda: SimpleNamespace(
            count_scan_success_as_processed=False,
            play_sound_for_duplicate_received_qr=False,
            offline_scan_mode=False,
        ),
        should_count_scan_success_as_processed=lambda settings=None: bool(
            getattr(settings, "count_scan_success_as_processed", False)
        ),
        should_play_sound_for_duplicate_received_qr=lambda settings=None: bool(
            getattr(settings, "play_sound_for_duplicate_received_qr", False)
        ),
    )
    return app


class AppPrintFlowTest(unittest.TestCase):
    def test_open_witchform_login_page_uses_configured_url_without_replacing_order_page(self) -> None:
        app = app_main.Application.__new__(app_main.Application)
        calls: list[tuple[str, bool]] = []
        app._browser_service = SimpleNamespace(
            login_url="https://example.com/login",
            open_page=lambda url, *, preserve_current_page: calls.append((url, preserve_current_page)) or True,
        )

        self.assertTrue(app.open_witchform_login_page())
        self.assertEqual(calls, [("https://example.com/login", True)])

    def test_success_flow_marks_prints_and_plays_sound(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)

        with (
            patch("main.print_order_receipt") as print_mock,
            patch.object(app, "_play_scan_success_sound") as sound_mock,
            patch.object(app, "_commit_scan_success_count") as commit_mock,
        ):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.complete_calls, 1)
        self.assertEqual(order_vm.mark_calls, 1)
        self.assertEqual(order_vm.rollback_calls, 0)
        print_mock.assert_called_once()
        sound_mock.assert_called_once_with(order, increment_count=True, persist_count=False)
        commit_mock.assert_called_once_with()
        self.assertEqual(len(app._excel_service.processing_time_calls), 1)
        self.assertEqual(app._excel_service.processing_time_calls[0][0], "ORDER-001")
        self.assertEqual(app._state, app_main.AppState.READY)

    def test_qr_scan_auto_print_off_marks_order_but_skips_print_and_rollback(self) -> None:
        order = Order(
            order_number="ORDER-001",
            name="홍길동",
            phone="010",
            seat="A-1",
            goods=[],
            order_status="거래중",
        )
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)
        app._receipt_settings = SimpleNamespace(
            qr_scan_success_sound_path="",
            qr_scan_auto_print_enabled=False,
        )
        app._settings_store = SimpleNamespace(load=lambda: app._receipt_settings)

        with (
            patch("main.print_order_receipt", side_effect=AssertionError("QR auto print should be skipped")) as print_mock,
            patch.object(app, "_play_scan_success_sound") as sound_mock,
            patch.object(app, "_commit_scan_success_count") as commit_mock,
        ):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.complete_calls, 1)
        self.assertEqual(order_vm.mark_calls, 1)
        self.assertEqual(order_vm.rollback_calls, 0)
        print_mock.assert_not_called()
        sound_mock.assert_called_once_with(order, increment_count=True, persist_count=False)
        commit_mock.assert_called_once_with()
        self.assertEqual(app._state, app_main.AppState.READY)
        self.assertIn("자동 출력 꺼짐", app._scanner_view.status_message)

    def test_online_success_records_received_at_and_processing_time(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)
        app._receipt_settings = SimpleNamespace(qr_scan_success_sound_path="", qr_scan_auto_print_enabled=False)
        app._settings_store = SimpleNamespace(load=lambda: app._receipt_settings)

        with patch("main.datetime", SimpleNamespace(now=lambda: datetime(2026, 8, 17, 12, 34, 56))):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order.received_at, "2026-08-17 12:34:56")
        self.assertEqual(app._excel_service.processing_time_calls, [("ORDER-001", "2026-08-17 12:34:56")])
        self.assertEqual(app._state, app_main.AppState.READY)

    def test_local_received_status_or_print_failure_never_writes_processing_time(self) -> None:
        for failure_stage in ("received", "status", "print"):
            with self.subTest(failure_stage=failure_stage):
                order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
                order_vm = _FakeOrderViewModel(order=order, mark_ok=failure_stage != "received")
                app = _build_app_with_order_vm(order_vm)
                app._excel_service = _FakeExcelService(status_ok=failure_stage != "status")

                with patch(
                    "main.print_order_receipt",
                    side_effect=RuntimeError("printer down") if failure_stage == "print" else None,
                ):
                    app._process_resolved_qr(
                        "https://witchform.com/qrcode_link.php?a=1",
                        BrowserResolveResult(ok=True, status_code=302, location="/x"),
                        allow_auth_retry=False,
                    )

                self.assertEqual(app._excel_service.processing_time_calls, [])
                self.assertEqual(app._state, app_main.AppState.ERROR)

    def test_online_success_reports_processing_time_writer_failure(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)
        app._excel_service = _FakeExcelService(processing_time_ok=False)
        app._receipt_settings = SimpleNamespace(qr_scan_success_sound_path="", qr_scan_auto_print_enabled=False)
        app._settings_store = SimpleNamespace(load=lambda: app._receipt_settings)

        with patch.object(app, "_commit_scan_success_count") as commit_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(len(app._excel_service.processing_time_calls), 1)
        self.assertEqual(app._excel_service.processing_time_calls[0][0], "ORDER-001")
        commit_mock.assert_not_called()
        self.assertEqual(app._state, app_main.AppState.ERROR)

    def test_print_failure_rolls_back_mark(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order, rollback_ok=True)
        app = _build_app_with_order_vm(order_vm)

        with patch("main.print_order_receipt", side_effect=RuntimeError("printer down")):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.mark_calls, 1)
        self.assertEqual(order_vm.rollback_calls, 1)
        self.assertEqual(app._excel_service.processing_time_calls, [])
        self.assertEqual(app._state, app_main.AppState.ERROR)
        self.assertIn("원복", app._scanner_view.status_message)

    def test_print_failure_restores_blank_app_owned_status_not_source_progress_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["주문번호", "주문자명", "주문상태", "진행상태"])
            worksheet.append(["ORDER-001", "홍길동", "", "결제완료"])
            workbook.save(file_path)
            workbook.close()

            excel_service = _TrackingExcelService(str(file_path))
            order = excel_service.find_order("ORDER-001")
            assert order is not None
            self.assertEqual(order.order_status, "결제완료")
            app = _build_app_with_order_vm(_FakeOrderViewModel(order=order))
            app._excel_service = excel_service

            with patch("main.print_order_receipt", side_effect=RuntimeError("printer down")):
                app._process_resolved_qr(
                    "https://witchform.com/qrcode_link.php?a=1",
                    BrowserResolveResult(ok=True, status_code=302, location="/x"),
                    allow_auth_retry=False,
                )

            loaded = load_workbook(file_path, read_only=True, data_only=True)
            try:
                self.assertIn(loaded.active["C2"].value, (None, ""))
                self.assertNotIn("처리시간", [cell.value for cell in loaded.active[1]])
            finally:
                loaded.close()
            self.assertEqual(excel_service.processing_time_calls, [])

    def test_received_order_skips_click_print_and_is_silent_by_default(self) -> None:
        order = Order(
            order_number="ORDER-001",
            name="홍길동",
            phone="010",
            seat="A-1",
            goods=[],
            received_at="2026-02-23 10:00:00",
        )
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)

        with patch("main.print_order_receipt") as print_mock, patch.object(app, "_play_scan_success_sound") as sound_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.open_calls, 0)
        self.assertEqual(order_vm.complete_calls, 0)
        self.assertEqual(order_vm.mark_calls, 0)
        print_mock.assert_not_called()
        sound_mock.assert_not_called()
        self.assertEqual(app._excel_service.processing_time_calls, [])
        self.assertEqual(app._state, app_main.AppState.READY)
        self.assertIs(app._order_view.last_order, order)
        self.assertIn("이미 수령완료", app._scanner_view.status_message)

    def _assert_cancelled_order_is_blocked(self, raw_status: str) -> None:
        order = Order(
            order_number="ORDER-001",
            name="홍길동",
            phone="010",
            seat="A-1",
            goods=[],
            order_status=raw_status,
        )
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)
        emitted_orders: list[Order] = []
        app._order_listener = emitted_orders.append
        app._state = app_main.AppState.PROCESSING
        app._scanner_view.scanning_enabled = False

        with (
            patch("main.print_order_receipt") as print_mock,
            patch.object(app, "_play_scan_success_sound") as sound_mock,
            patch.object(app, "_commit_scan_success_count") as commit_mock,
        ):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.load_calls, 1)
        self.assertIsNone(app._order_view.last_order)
        self.assertEqual(emitted_orders, [])
        self.assertEqual(order_vm.open_calls, 0)
        self.assertEqual(order_vm.complete_calls, 0)
        self.assertEqual(order_vm.mark_calls, 0)
        self.assertEqual(order_vm.rollback_calls, 0)
        self.assertEqual(app._excel_service.status_calls, [])
        self.assertEqual(app._excel_service.processing_time_calls, [])
        print_mock.assert_not_called()
        sound_mock.assert_not_called()
        commit_mock.assert_not_called()
        self.assertEqual(app._state, app_main.AppState.READY)
        self.assertTrue(app._scanner_view.scanning_enabled)
        self.assertIn("취소", app._scanner_view.status_message)
        self.assertNotIn(order.order_number, app._scanner_view.status_message)

    def test_cancelled_order_status_blocks_before_order_detail_emit_and_success_side_effects(self) -> None:
        for raw_status in ("주문취소", " 주문취소 "):
            with self.subTest(raw_status=raw_status):
                self._assert_cancelled_order_is_blocked(raw_status)

    def test_auto_cancelled_order_status_blocks_before_order_detail_emit_and_success_side_effects(self) -> None:
        for raw_status in ("자동주문취소", " 자동주문취소 "):
            with self.subTest(raw_status=raw_status):
                self._assert_cancelled_order_is_blocked(raw_status)

    def test_load_order_missing_reports_error_without_side_effects(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order, load_result=None)
        app = _build_app_with_order_vm(order_vm)

        with patch("main.print_order_receipt") as print_mock, patch.object(app, "_play_scan_success_sound") as sound_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(app._state, app_main.AppState.ERROR)
        self.assertEqual(order_vm.open_calls, 0)
        self.assertEqual(order_vm.complete_calls, 0)
        self.assertEqual(order_vm.mark_calls, 0)
        print_mock.assert_not_called()
        sound_mock.assert_not_called()
        self.assertEqual(app._excel_service.processing_time_calls, [])
        self.assertIn("주문번호를 찾을 수 없습니다", app._scanner_view.status_message)

    def test_already_received_click_result_marks_order_and_is_silent_by_default(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(
            order=order,
            click_result=ReceiptClickResult(
                success=False,
                error_code="ALREADY_RECEIVED",
                error_message="이미 수령완료 처리된 주문입니다.",
            ),
        )
        app = _build_app_with_order_vm(order_vm)

        with patch.object(app, "_play_scan_success_sound") as sound_mock, patch("main.print_order_receipt") as print_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.complete_calls, 1)
        self.assertEqual(order_vm.mark_calls, 1)
        print_mock.assert_not_called()
        sound_mock.assert_not_called()
        self.assertEqual(app._excel_service.processing_time_calls, [])
        self.assertEqual(app._state, app_main.AppState.READY)
        self.assertIn("이미 수령완료", app._scanner_view.status_message)

    def test_open_current_order_page_failure_reports_error_without_receipt_side_effects(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order, open_ok=False)
        app = _build_app_with_order_vm(order_vm)

        with patch("main.print_order_receipt") as print_mock, patch.object(app, "_play_scan_success_sound") as sound_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(app._state, app_main.AppState.ERROR)
        self.assertEqual(order_vm.open_calls, 1)
        self.assertEqual(order_vm.complete_calls, 0)
        self.assertEqual(order_vm.mark_calls, 0)
        print_mock.assert_not_called()
        sound_mock.assert_not_called()
        self.assertEqual(app._excel_service.processing_time_calls, [])
        self.assertEqual(app._scanner_view.status_message, "주문 상세 페이지를 열 수 없습니다.")

    def test_open_current_order_page_exception_reports_recoverable_error_without_stopping_loop(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(
            order=order,
            open_exception=RuntimeError("작업 시간 초과: open_page"),
        )
        app = _build_app_with_order_vm(order_vm)

        with (
            patch("main.print_order_receipt") as print_mock,
            patch.object(app, "_play_scan_success_sound") as sound_mock,
        ):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(app._state, app_main.AppState.ERROR)
        self.assertEqual(order_vm.open_calls, 1)
        self.assertEqual(order_vm.complete_calls, 0)
        self.assertEqual(order_vm.mark_calls, 0)
        print_mock.assert_not_called()
        sound_mock.assert_not_called()
        self.assertEqual(app._excel_service.processing_time_calls, [])
        self.assertEqual(app._scanner_view.status_message, "주문 상세 페이지를 열 수 없습니다.")

    def test_click_failure_reports_generic_receipt_failure_status_message(self) -> None:
        order = Order(order_number="ORDER-001", name="TEST USER", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(
            order=order,
            click_result=ReceiptClickResult(
                success=False,
                error_code="PRIMARY_CLICK_FAIL",
                error_message="수령 완료 1차 버튼 클릭 실패: timeout",
            ),
        )
        app = _build_app_with_order_vm(order_vm)

        with patch.object(app, "_play_scan_success_sound") as sound_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(app._state, app_main.AppState.ERROR)
        self.assertIn("수령 완료 처리 상태를 확인하지 못했습니다", app._scanner_view.status_message)
        self.assertEqual(app._excel_service.processing_time_calls, [])
        sound_mock.assert_called_once_with(order, increment_count=False)

    def test_mark_current_order_received_failure_stops_before_print_and_sound(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order, mark_ok=False)
        app = _build_app_with_order_vm(order_vm)

        with patch("main.print_order_receipt") as print_mock, patch.object(app, "_play_scan_success_sound") as sound_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(app._state, app_main.AppState.ERROR)
        self.assertEqual(order_vm.complete_calls, 1)
        self.assertEqual(order_vm.mark_calls, 1)
        print_mock.assert_not_called()
        sound_mock.assert_not_called()
        self.assertEqual(app._excel_service.processing_time_calls, [])
        self.assertIn("수령확인 저장 실패", app._scanner_view.status_message)

    def test_browser_failure_uses_user_friendly_status_message(self) -> None:
        order = Order(order_number="ORDER-001", name="TEST USER", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)

        app._process_resolved_qr(
            "https://witchform.com/qrcode_link.php?a=1",
            BrowserResolveResult(
                ok=False,
                error_code="NETWORK_ERROR",
                error_message="브라우저 요청 실패: APIRequestContext.get: Target page closed",
            ),
            allow_auth_retry=False,
        )

        self.assertEqual(app._state, app_main.AppState.ERROR)
        self.assertEqual(
            app._scanner_view.status_message,
            "브라우저 요청을 확인하지 못했습니다. 다시 스캔해주세요.",
        )
        self.assertEqual(app._excel_service.processing_time_calls, [])

    def test_qr_validation_parse_and_browser_request_failures_never_write_processing_time(self) -> None:
        cases = (
            ("invalid_url", "not-a-url", None),
            ("invalid_prefix", "https://example.com/qrcode_link.php?a=1", None),
            ("browser_exception", "https://witchform.com/qrcode_link.php?a=1", RuntimeError("browser down")),
        )
        for failure_stage, qr_url, browser_exception in cases:
            with self.subTest(failure_stage=failure_stage):
                order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
                app = _build_app_with_order_vm(_FakeOrderViewModel(order=order))
                if browser_exception is not None:
                    app._browser_service = SimpleNamespace(
                        resolve_qr_redirect=lambda _url: (_ for _ in ()).throw(browser_exception)
                    )

                app._process_qr(qr_url, allow_auth_retry=False)

                self.assertEqual(app._excel_service.processing_time_calls, [])
                self.assertEqual(app._state, app_main.AppState.ERROR)

        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        app = _build_app_with_order_vm(_FakeOrderViewModel(order=order))
        app._api_service = SimpleNamespace(
            parse_qr_redirect=lambda *_args: QRParseResult(success=False, error_code="INVALID_REDIRECT", error_message="잘못된 QR")
        )

        app._process_resolved_qr(
            "https://witchform.com/qrcode_link.php?a=1",
            BrowserResolveResult(ok=True, status_code=302, location="/x"),
            allow_auth_retry=False,
        )

        self.assertEqual(app._excel_service.processing_time_calls, [])
        self.assertEqual(app._state, app_main.AppState.ERROR)

    def test_success_flow_keeps_working_without_order_window(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)
        app._order_view = None

        with patch("main.print_order_receipt") as print_mock, patch.object(app, "_play_scan_success_sound") as sound_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(app._state, app_main.AppState.READY)
        self.assertEqual(order_vm.complete_calls, 1)
        self.assertEqual(order_vm.mark_calls, 1)
        print_mock.assert_called_once()
        sound_mock.assert_called_once_with(order, increment_count=True, persist_count=False)

    def test_listener_failures_are_logged_without_breaking_success_flow(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)

        def _broken_order_listener(_order: Order) -> None:
            raise RuntimeError("order callback boom")

        def _broken_status_listener(_state: str, _message: str) -> None:
            raise RuntimeError("status callback boom")

        app._order_listener = _broken_order_listener
        app._status_listener = _broken_status_listener

        with (
            self.assertLogs("main", level="WARNING") as captured,
            patch("main.print_order_receipt") as print_mock,
            patch.object(app, "_play_scan_success_sound") as sound_mock,
        ):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(app._state, app_main.AppState.READY)
        print_mock.assert_called_once()
        sound_mock.assert_called_once_with(order, increment_count=True, persist_count=False)
        self.assertTrue(any("주문 콜백 처리 실패" in line for line in captured.output))
        self.assertTrue(any("상태 콜백 처리 실패" in line for line in captured.output))

    def test_network_timeout_retries_once_and_reprocesses_success(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)
        retry_result = BrowserResolveResult(ok=True, status_code=302, location="/orders/1")
        reprocessed: list[tuple[str, BrowserResolveResult, bool]] = []
        errors: list[str] = []

        app._browser_service = SimpleNamespace(resolve_qr_redirect=lambda _qr_url: retry_result)
        app._process_resolved_qr = lambda qr_url, browser_result, allow_auth_retry: reprocessed.append(
            (qr_url, browser_result, allow_auth_retry)
        )
        app._enter_error = lambda message, keep_auth_state=True: errors.append(message)

        app._handle_browser_failure(
            BrowserResolveResult(
                ok=False,
                error_code="NETWORK_TIMEOUT",
                error_message="timeout",
            ),
            "https://witchform.com/qrcode_link.php?a=1",
            allow_auth_retry=True,
        )

        self.assertEqual(errors, [])
        self.assertEqual(
            reprocessed,
            [("https://witchform.com/qrcode_link.php?a=1", retry_result, False)],
        )

    def test_network_timeout_reports_fixed_error_when_retry_also_fails(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)
        errors: list[str] = []

        app._browser_service = SimpleNamespace(
            resolve_qr_redirect=lambda _qr_url: BrowserResolveResult(
                ok=False,
                error_code="NETWORK_TIMEOUT",
                error_message="timeout",
            )
        )
        app._enter_error = lambda message, keep_auth_state=True: errors.append(message)

        app._handle_browser_failure(
            BrowserResolveResult(
                ok=False,
                error_code="NETWORK_TIMEOUT",
                error_message="timeout",
            ),
            "https://witchform.com/qrcode_link.php?a=1",
            allow_auth_retry=True,
        )

        self.assertEqual(errors, ["네트워크 시간 초과가 발생했습니다"])


    def test_received_order_can_play_general_sound_when_debug_duplicate_sound_enabled(self) -> None:
        order = Order(
            order_number="ORDER-001",
            name="TEST USER",
            phone="010",
            seat="A-1",
            goods=[],
            received_at="2026-02-23 10:00:00",
        )
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)
        app._ticket_debug_tools_service = SimpleNamespace(
            load_settings=lambda: SimpleNamespace(
                count_scan_success_as_processed=False,
                play_sound_for_duplicate_received_qr=True,
                offline_scan_mode=False,
            ),
            should_play_sound_for_duplicate_received_qr=lambda settings=None: bool(
                getattr(settings, "play_sound_for_duplicate_received_qr", False)
            ),
        )

        with patch("main.print_order_receipt") as print_mock, patch.object(app, "_play_scan_success_sound") as sound_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.open_calls, 0)
        self.assertEqual(order_vm.complete_calls, 0)
        self.assertEqual(order_vm.mark_calls, 0)
        print_mock.assert_not_called()
        sound_mock.assert_called_once_with(order, increment_count=False, persist_count=True)
        self.assertEqual(app._state, app_main.AppState.READY)

    def test_received_order_can_increment_count_and_play_counted_sound_when_debug_count_enabled(self) -> None:
        order = Order(
            order_number="ORDER-001",
            name="TEST USER",
            phone="010",
            seat="A-1",
            goods=[],
            received_at="2026-02-23 10:00:00",
        )
        order_vm = _FakeOrderViewModel(order=order)
        app = _build_app_with_order_vm(order_vm)
        app._ticket_debug_tools_service = SimpleNamespace(
            load_settings=lambda: SimpleNamespace(
                count_scan_success_as_processed=True,
                play_sound_for_duplicate_received_qr=False,
                offline_scan_mode=False,
            ),
            should_play_sound_for_duplicate_received_qr=lambda settings=None: bool(
                getattr(settings, "play_sound_for_duplicate_received_qr", False)
            ),
            should_count_scan_success_as_processed=lambda settings=None: bool(
                getattr(settings, "count_scan_success_as_processed", False)
            ),
        )

        with patch.object(app, "_commit_scan_success_count") as commit_mock, patch.object(
            app, "_play_scan_success_sound"
        ) as sound_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        sound_mock.assert_called_once_with(order, increment_count=True, persist_count=False)
        commit_mock.assert_called_once_with()
        self.assertEqual(app._state, app_main.AppState.READY)

    def test_already_received_click_result_can_play_general_sound_when_debug_duplicate_sound_enabled(self) -> None:
        order = Order(order_number="ORDER-001", name="TEST USER", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(
            order=order,
            click_result=ReceiptClickResult(
                success=False,
                error_code="ALREADY_RECEIVED",
                error_message="?대? ?섎졊?꾨즺 泥섎━??二쇰Ц?낅땲??",
            ),
        )
        app = _build_app_with_order_vm(order_vm)
        app._ticket_debug_tools_service = SimpleNamespace(
            load_settings=lambda: SimpleNamespace(
                count_scan_success_as_processed=False,
                play_sound_for_duplicate_received_qr=True,
                offline_scan_mode=False,
            ),
            should_play_sound_for_duplicate_received_qr=lambda settings=None: bool(
                getattr(settings, "play_sound_for_duplicate_received_qr", False)
            ),
        )

        with patch.object(app, "_play_scan_success_sound") as sound_mock, patch("main.print_order_receipt") as print_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.complete_calls, 1)
        self.assertEqual(order_vm.mark_calls, 1)
        print_mock.assert_not_called()
        sound_mock.assert_called_once_with(order, increment_count=False, persist_count=True)
        self.assertEqual(app._state, app_main.AppState.READY)

    def test_already_received_click_result_can_increment_count_and_play_counted_sound_when_debug_count_enabled(self) -> None:
        order = Order(order_number="ORDER-001", name="TEST USER", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(
            order=order,
            click_result=ReceiptClickResult(
                success=False,
                error_code="ALREADY_RECEIVED",
                error_message="이미 수령완료 처리된 주문입니다.",
            ),
        )
        app = _build_app_with_order_vm(order_vm)
        app._ticket_debug_tools_service = SimpleNamespace(
            load_settings=lambda: SimpleNamespace(
                count_scan_success_as_processed=True,
                play_sound_for_duplicate_received_qr=False,
                offline_scan_mode=False,
            ),
            should_play_sound_for_duplicate_received_qr=lambda settings=None: bool(
                getattr(settings, "play_sound_for_duplicate_received_qr", False)
            ),
            should_count_scan_success_as_processed=lambda settings=None: bool(
                getattr(settings, "count_scan_success_as_processed", False)
            ),
        )

        with patch.object(app, "_commit_scan_success_count") as commit_mock, patch.object(
            app, "_play_scan_success_sound"
        ) as sound_mock, patch("main.print_order_receipt") as print_mock:
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        print_mock.assert_not_called()
        sound_mock.assert_called_once_with(order, increment_count=True, persist_count=False)
        commit_mock.assert_called_once_with()
        self.assertEqual(app._state, app_main.AppState.READY)


class ReceiptClickRetryTest(unittest.TestCase):
    """complete_receipt() 자동 재시도 및 주문상태 롤백 시나리오 검증."""

    def _make_app(self, order_vm: _FakeOrderViewModel):
        app = _build_app_with_order_vm(order_vm)
        app._excel_service = _FakeExcelService()
        return app

    def test_click_fail_then_retry_succeeds_completes_normally(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(
            order=order,
            click_result_seq=[
                ReceiptClickResult(success=False, error_code="PRIMARY_CLICK_FAIL", error_message="로딩 지연"),
                ReceiptClickResult(success=True),
            ],
        )
        app = self._make_app(order_vm)

        with patch("main.print_order_receipt"), patch("main.time") as mock_time:
            mock_time.sleep = lambda _s: None
            mock_time.monotonic = __import__("time").monotonic
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.complete_calls, 2)
        self.assertEqual(app._state, app_main.AppState.READY)

    def test_click_fail_then_retry_also_fails_enters_error(self) -> None:
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(
            order=order,
            click_result_seq=[
                ReceiptClickResult(success=False, error_code="PRIMARY_CLICK_FAIL", error_message="실패1"),
                ReceiptClickResult(success=False, error_code="PRIMARY_CLICK_FAIL", error_message="실패2"),
            ],
        )
        app = self._make_app(order_vm)

        with patch("main.print_order_receipt"), patch("main.time") as mock_time:
            mock_time.sleep = lambda _s: None
            mock_time.monotonic = __import__("time").monotonic
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.complete_calls, 2)
        self.assertEqual(app._state, app_main.AppState.ERROR)

    def test_print_failure_without_raw_status_source_does_not_write_display_status(self) -> None:
        order = Order(
            order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[],
            order_status="거래중",
        )
        order_vm = _FakeOrderViewModel(order=order)
        app = self._make_app(order_vm)

        status_calls: list[tuple[str, str]] = []
        app._excel_service = SimpleNamespace(
            mark_order_status=lambda on, st: status_calls.append((on, st)) or True,
            mark_order_processing_time=lambda *_args: True,
        )

        with patch("main.print_order_receipt", side_effect=RuntimeError("프린터 오류")):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

            self.assertEqual(app._state, app_main.AppState.ERROR)
            self.assertEqual(status_calls, [("ORDER-001", "거래종료")])


class SessionTimeoutRecoveryTest(unittest.TestCase):
    """_recover_auth_and_retry() 세션 타임아웃 후 READY 복구 검증."""

    def test_login_timeout_returns_to_ready_state(self) -> None:
        app = app_main.Application.__new__(app_main.Application)
        app._state = app_main.AppState.RECOVERING
        app._scanner_view = _FakeScannerView()
        app._stop_requested = False

        with patch.object(app, "_wait_for_login", return_value=False), \
             patch.object(app, "_is_stop_requested", return_value=False):
            app._recover_auth_and_retry("https://witchform.com/qrcode_link.php?a=1")

        self.assertEqual(app._state, app_main.AppState.READY)


class ResolveQrOfflineTest(unittest.TestCase):
    """_resolve_qr_offline: test_order 단축 경로와 httpx 경로를 분리 검증."""

    def _make_app(self) -> app_main.Application:
        app = app_main.Application.__new__(app_main.Application)
        return app

    def test_test_order_param_returns_fake_redirect_without_http(self) -> None:
        app = self._make_app()
        url = "https://witchform.com/qrcode_link.php?test_order=WFLM_123456"
        with patch("httpx.get") as mock_get:
            result = app._resolve_qr_offline(url)
        mock_get.assert_not_called()
        self.assertTrue(result.ok)
        self.assertEqual(result.status_code, 302)
        self.assertEqual(result.location, "/w/myform/sellForm-history-detail/WFLM_123456")

    def test_test_order_param_is_uppercased_in_location(self) -> None:
        app = self._make_app()
        url = "https://witchform.com/qrcode_link.php?test_order=wflm_abcdef"
        with patch("httpx.get") as mock_get:
            result = app._resolve_qr_offline(url)
        mock_get.assert_not_called()
        self.assertIn("wflm_abcdef", result.location)

    def test_no_test_order_param_calls_httpx_with_follow_redirects(self) -> None:
        app = self._make_app()
        url = "https://witchform.com/qrcode_link.php?a=1"
        final_url = "https://witchform.com/w/myform/sellForm-history-detail/ORDER-001"
        fake_resp = SimpleNamespace(
            status_code=200,
            url=final_url,
        )
        with patch("httpx.get", return_value=fake_resp) as mock_get:
            result = app._resolve_qr_offline(url)
        _, kwargs = mock_get.call_args
        self.assertTrue(kwargs.get("follow_redirects", False), "follow_redirects=True 이어야 합니다")
        self.assertTrue(result.ok)
        self.assertEqual(result.location, final_url)

    def test_login_final_url_returns_fake_302(self) -> None:
        app = self._make_app()
        url = "https://witchform.com/qrcode_link.php?a=1"
        login_url = "https://witchform.com/w/login?redirect=%2Fw%2Fmyform%2FsellForm-history-detail%2FORDER-001"
        fake_resp = SimpleNamespace(
            status_code=200,
            url=login_url,
        )
        with patch("httpx.get", return_value=fake_resp):
            result = app._resolve_qr_offline(url)
        self.assertTrue(result.ok)
        self.assertEqual(result.status_code, 302)
        self.assertEqual(result.location, login_url)

    def test_httpx_exception_returns_error_result(self) -> None:
        app = self._make_app()
        url = "https://witchform.com/qrcode_link.php?a=1"
        with patch("httpx.get", side_effect=Exception("연결 실패")):
            result = app._resolve_qr_offline(url)
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, "OFFLINE_HTTP_FAIL")


class OnlineModeGuaranteeTest(unittest.TestCase):
    """온라인 모드에서 complete_receipt 성공 없이는 수령완료/카운트가 절대 실행되지 않음을 보호한다."""

    def _make_app(self, order_vm: _FakeOrderViewModel):
        app = _build_app_with_order_vm(order_vm)
        app._excel_service = _FakeExcelService()
        return app

    def test_click_failure_never_reaches_main_mark_and_count(self) -> None:
        """일반 실패 에러 코드에서 메인 경로의 수령완료 표시와 카운트가 호출되지 않는다."""
        failure_codes = [
            "PAGE_NOT_FOUND",
            "TIMEOUT",
            "RECEIPT_EXCEPTION",
            "UNKNOWN_FAILURE",
        ]
        for error_code in failure_codes:
            with self.subTest(error_code=error_code):
                order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
                order_vm = _FakeOrderViewModel(
                    order=order,
                    click_result=ReceiptClickResult(success=False, error_code=error_code),
                )
                app = self._make_app(order_vm)
                with (
                    patch("main.print_order_receipt") as print_mock,
                    patch.object(app, "_commit_scan_success_count") as commit_mock,
                ):
                    app._process_resolved_qr(
                        "https://witchform.com/qrcode_link.php?a=1",
                        BrowserResolveResult(ok=True, status_code=302, location="/x"),
                        allow_auth_retry=False,
                    )
                self.assertEqual(order_vm.mark_calls, 0, f"mark_calls != 0 for error_code={error_code}")
                commit_mock.assert_not_called()
                print_mock.assert_not_called()

    def test_click_retry_failures_never_reach_main_mark_and_count(self) -> None:
        """재시도 포함 실패 경로(PRIMARY/CONFIRM_CLICK_FAIL)에서도 mark/count가 호출되지 않는다."""
        for error_code in ("PRIMARY_CLICK_FAIL", "CONFIRM_CLICK_FAIL"):
            with self.subTest(error_code=error_code):
                order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
                order_vm = _FakeOrderViewModel(
                    order=order,
                    click_result_seq=[
                        ReceiptClickResult(success=False, error_code=error_code),
                        ReceiptClickResult(success=False, error_code=error_code),
                    ],
                )
                app = self._make_app(order_vm)
                with (
                    patch("main.print_order_receipt") as print_mock,
                    patch.object(app, "_commit_scan_success_count") as commit_mock,
                    patch("main.time") as mock_time,
                ):
                    mock_time.sleep = lambda _s: None
                    mock_time.monotonic = __import__("time").monotonic
                    app._process_resolved_qr(
                        "https://witchform.com/qrcode_link.php?a=1",
                        BrowserResolveResult(ok=True, status_code=302, location="/x"),
                        allow_auth_retry=False,
                    )
                self.assertEqual(order_vm.mark_calls, 0, f"mark_calls != 0 for error_code={error_code}")
                commit_mock.assert_not_called()
                print_mock.assert_not_called()

    def test_complete_receipt_exception_blocks_mark_and_count(self) -> None:
        """complete_receipt가 예외를 던지면 수령완료 표시와 카운트가 실행되지 않는다."""
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = self._make_app(order_vm)

        def _raise() -> ReceiptClickResult:
            raise RuntimeError("브라우저 예외 발생")

        order_vm.complete_receipt = _raise  # type: ignore[assignment]

        with (
            patch("main.print_order_receipt") as print_mock,
            patch.object(app, "_commit_scan_success_count") as commit_mock,
        ):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        self.assertEqual(order_vm.mark_calls, 0)
        commit_mock.assert_not_called()
        print_mock.assert_not_called()
        self.assertEqual(app._state, app_main.AppState.ERROR)

    def test_count_not_committed_when_print_fails(self) -> None:
        """영수증 출력이 실패하면 카운트가 확정되지 않는다."""
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = self._make_app(order_vm)

        with (
            patch("main.print_order_receipt", side_effect=RuntimeError("프린터 오류")),
            patch.object(app, "_commit_scan_success_count") as commit_mock,
        ):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        commit_mock.assert_not_called()
        self.assertEqual(app._state, app_main.AppState.ERROR)

    def test_count_committed_only_after_full_success(self) -> None:
        """complete_receipt + mark + print 모두 성공한 후에만 카운트가 확정된다."""
        order = Order(order_number="ORDER-001", name="홍길동", phone="010", seat="A-1", goods=[])
        order_vm = _FakeOrderViewModel(order=order)
        app = self._make_app(order_vm)

        with (
            patch("main.print_order_receipt"),
            patch.object(app, "_commit_scan_success_count") as commit_mock,
        ):
            app._process_resolved_qr(
                "https://witchform.com/qrcode_link.php?a=1",
                BrowserResolveResult(ok=True, status_code=302, location="/x"),
                allow_auth_retry=False,
            )

        commit_mock.assert_called_once()
        self.assertEqual(order_vm.complete_calls, 1)
        self.assertEqual(order_vm.mark_calls, 1)
        self.assertEqual(app._state, app_main.AppState.READY)


if __name__ == "__main__":
    unittest.main()
