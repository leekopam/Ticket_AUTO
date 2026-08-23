"""엑셀 수령확인 상태 관리 테스트."""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from services.excel_service import (
    ExcelService,
    ORDER_STATUS_HEADER,
    PROCESSING_TIME_HEADER,
    RECEIPT_HEADER,
    SOURCE_PROGRESS_STATUS_HEADER,
)


def _create_workbook(path: Path, *, header_order_label: str = "주문번호") -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "번호",
            header_order_label,
            "주문자명",
            "주문자연락처",
            "좌석번호",
            "[상품1] 티켓",
        ]
    )
    ws.append([1, "ORDER-001", "홍길동", "010-0000-0000", "A-1", 2])
    wb.save(path)
    wb.close()


class ExcelReceiptStateTest(unittest.TestCase):
    def test_mark_order_status_does_not_fill_search_receipt_time(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            _create_workbook(file_path)
            service = ExcelService(str(file_path))

            self.assertTrue(service.mark_order_status("ORDER-001", "거래종료"))

            loaded = load_workbook(file_path, read_only=True, data_only=True)
            ws = loaded.active
            headers = [cell.value for cell in ws[1]]
            status_col = headers.index(ORDER_STATUS_HEADER) + 1
            self.assertEqual(ws.cell(row=2, column=status_col).value, "거래종료")
            loaded.close()

            from views import dashboard_flet_view as dashboard

            view_state = dashboard.build_order_search_view_state(
                "",
                "전체",
                service.search_orders(),
                [],
                None,
            )
            self.assertEqual(view_state.row_states[0].order_status_text, "")

    def test_mark_order_received_creates_header_and_persists_value(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            _create_workbook(file_path)
            service = ExcelService(str(file_path))

            ok = service.mark_order_received("ORDER-001", "2026-02-23 11:22:33")
            self.assertTrue(ok)

            loaded = load_workbook(file_path)
            ws = loaded.active
            headers = [cell.value for cell in ws[1]]
            self.assertIn(RECEIPT_HEADER, headers)
            receipt_col = headers.index(RECEIPT_HEADER) + 1
            self.assertEqual(ws.cell(row=2, column=receipt_col).value, "2026-02-23 11:22:33")
            loaded.close()

            order = service.find_order("ORDER-001")
            self.assertIsNotNone(order)
            assert order is not None
            self.assertEqual(order.received_at, "2026-02-23 11:22:33")
            self.assertTrue(order.is_received)

    def test_rollback_order_received_restores_previous_value(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            _create_workbook(file_path)
            service = ExcelService(str(file_path))

            self.assertTrue(service.mark_order_received("ORDER-001", "2026-02-23 09:00:00"))
            self.assertTrue(service.rollback_order_received("ORDER-001", ""))

            order = service.find_order("ORDER-001")
            self.assertIsNotNone(order)
            assert order is not None
            self.assertEqual(order.received_at, "")
            self.assertFalse(order.is_received)

    def test_get_received_status_map_returns_empty_for_missing_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            missing_path = Path(temp_dir) / "missing.xlsx"
            service = ExcelService(str(missing_path))

            self.assertEqual(service.get_received_status_map(), {})

    def test_get_received_status_map_includes_only_received_orders(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["주문번호", RECEIPT_HEADER])
            ws.append(["ORDER-001", "2026-02-23 09:00:00"])
            ws.append(["ORDER-002", ""])
            ws.append(["", "2026-02-23 10:00:00"])
            ws.append(["ORDER-003", "2026-02-23 11:00:00"])
            wb.save(file_path)
            wb.close()
            service = ExcelService(str(file_path))

            self.assertEqual(
                service.get_received_status_map(),
                {
                    "ORDER-001": "2026-02-23 09:00:00",
                    "ORDER-003": "2026-02-23 11:00:00",
                },
            )

    def test_header_spacing_variation_still_matches_order_column(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            _create_workbook(file_path, header_order_label=" 주문번호 ")
            service = ExcelService(str(file_path))

            self.assertTrue(service.mark_order_received("ORDER-001", "2026-02-23 12:00:00"))
            order = service.find_order("ORDER-001")
            self.assertIsNotNone(order)
            assert order is not None
            self.assertEqual(order.received_at, "2026-02-23 12:00:00")

    def test_order_status_prefers_app_value_then_falls_back_to_source_progress(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["주문번호", "주문자명", ORDER_STATUS_HEADER, SOURCE_PROGRESS_STATUS_HEADER])
            worksheet.append(["ORDER-001", "A", "거래종료", "결제완료"])
            worksheet.append(["ORDER-002", "B", "", "결제완료"])
            worksheet.append(["ORDER-003", "C", "", ""])
            workbook.save(file_path)
            workbook.close()

            service = ExcelService(str(file_path))
            orders = service.search_orders()

            self.assertEqual([order.order_status for order in orders], ["거래종료", "결제완료", ""])
            self.assertEqual(service.find_order("ORDER-002").order_status, "결제완료")
            self.assertEqual(service.find_orders_by_customer(name="B")[0].order_status, "결제완료")

            from views import dashboard_flet_view as dashboard

            view_state = dashboard.build_order_search_view_state("", "전체", orders, [], None)
            self.assertEqual([row.order_number for row in view_state.row_states], ["ORDER-001", "ORDER-002"])
            self.assertEqual([row.order_status_text for row in view_state.row_states], ["", ""])
            self.assertEqual(dashboard.build_search_result_row_state(orders[2], [], 0).order_status_text, "")

    def test_search_status_cell_uses_received_at_without_legacy_processing_time_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append([
                "주문번호",
                "주문자명",
                ORDER_STATUS_HEADER,
                SOURCE_PROGRESS_STATUS_HEADER,
                RECEIPT_HEADER,
                PROCESSING_TIME_HEADER,
            ])
            worksheet.append(["ORDER-001", "A", "거래종료", "결제완료", "", "2026-02-23 09:00:00"])
            worksheet.append(["ORDER-002", "B", "", "결제완료", "", ""])
            worksheet.append(["ORDER-003", "C", "", "", "", "2026-02-23 10:00:00"])
            worksheet.append(["ORDER-004", "D", "거래종료", "결제완료", "2026-02-23 11:00:00", ""])
            workbook.save(file_path)
            workbook.close()

            service = ExcelService(str(file_path))
            orders = {order.order_number: order for order in service.search_orders()}

            from views import dashboard_flet_view as dashboard

            row_texts = {
                order_number: dashboard.build_search_result_row_state(order, [], 0).order_status_text
                for order_number, order in orders.items()
            }

            self.assertEqual(
                row_texts,
                {
                    "ORDER-001": "",
                    "ORDER-002": "",
                    "ORDER-003": "",
                    "ORDER-004": "2026-02-23 11:00:00",
                },
            )

    def test_processing_time_header_is_single_final_column_and_preserves_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["주문번호", PROCESSING_TIME_HEADER, "주문자명", PROCESSING_TIME_HEADER, "비고"])
            worksheet.append(["ORDER-001", "", "A", "2026-02-23 09:00:00", "first"])
            worksheet.append(["ORDER-002", "2026-02-23 10:00:00", "B", "", "second"])
            workbook.save(file_path)
            workbook.close()

            self.assertTrue(ExcelService(str(file_path)).ensure_processing_time_column())

            loaded = load_workbook(file_path, data_only=True)
            worksheet = loaded.active
            self.assertEqual([cell.value for cell in worksheet[1]], ["주문번호", "주문자명", "비고", PROCESSING_TIME_HEADER])
            self.assertEqual(list(worksheet.iter_rows(min_row=2, values_only=True)), [
                ("ORDER-001", "A", "first", "2026-02-23 09:00:00"),
                ("ORDER-002", "B", "second", "2026-02-23 10:00:00"),
            ])
            loaded.close()

    def test_processing_time_header_is_added_last_when_absent_and_updated_by_order(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "data.xlsx"
            _create_workbook(file_path)
            service = ExcelService(str(file_path))

            self.assertTrue(service.ensure_processing_time_column())
            self.assertTrue(service.mark_order_processing_time("ORDER-001", "2026-02-23 11:22:33"))

            loaded = load_workbook(file_path, data_only=True)
            worksheet = loaded.active
            self.assertEqual(worksheet.cell(row=1, column=worksheet.max_column).value, PROCESSING_TIME_HEADER)
            self.assertEqual(worksheet.cell(row=2, column=worksheet.max_column).value, "2026-02-23 11:22:33")
            loaded.close()


if __name__ == "__main__":
    unittest.main()
