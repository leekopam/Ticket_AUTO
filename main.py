"""QR 스캐너 기반 주문 수령 자동화 애플리케이션."""
from __future__ import annotations

import logging
import math
import threading
import time
from datetime import datetime
from enum import Enum, auto
from pathlib import Path
from typing import Callable
from urllib.parse import urlparse

from openpyxl import load_workbook

from models.order_model import Order
from models.receipt_settings_model import ReceiptSettings
from models.ticket_debug_settings_model import TicketDebugSettings
from project_paths import ensure_managed_data_file, resolve_project_path
from services.api_service import ApiService, QRParseResult
from services.browser_service import BrowserResolveResult, BrowserService, ReceiptClickResult
from services.excel_service import ExcelService
from services.receipt_print_pipeline import print_order_receipt
from services.receipt_settings_store import ReceiptSettingsStore
from services.scan_success_sound_service import ScanSuccessSoundService
from services.ticket_debug_tools_service import TicketDebugToolsService
from services.windows_audio_service import WindowsAudioService
from viewmodels.order_viewmodel import OrderViewModel
from views.order_view import OrderView
from views.scanner_view import ScannerView

logger = logging.getLogger(__name__)
PROJECT_ROOT = Path(__file__).resolve().parent


def _mask_name(value: str) -> str:
    """로그에 남길 때 이름 일부만 노출한다."""
    text = (value or "").strip()
    if not text:
        return ""
    if len(text) == 1:
        return "*"
    if len(text) == 2:
        return f"{text[0]}*"
    return f"{text[0]}*{text[-1]}"


def _mask_phone(value: str) -> str:
    """로그에 남길 때 연락처 중간 자리를 마스킹한다."""
    text = (value or "").strip()
    if not text:
        return ""
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) == 11:
        return f"{digits[:3]}-****-{digits[-4:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-***-{digits[-4:]}"
    if len(text) <= 4:
        return "****"
    return f"{text[:2]}***{text[-2:]}"


class AppState(Enum):
    """애플리케이션 동작 상태."""

    AUTH_WAIT = auto()
    READY = auto()
    PROCESSING = auto()
    RECOVERING = auto()
    ERROR = auto()


StatusListener = Callable[[str, str], None]


class Application:
    """주문 수령 자동화 애플리케이션."""

    def __init__(self, *, show_order_window: bool = True, camera_index: int = 0):
        self._state = AppState.AUTH_WAIT
        self._show_order_window = bool(show_order_window)

        self._excel_service = ExcelService(str(ensure_managed_data_file()))
        self._excel_service.ensure_seat_column()
        self._excel_service.ensure_receipt_column()
        self._excel_service.ensure_order_status_column()
        self._excel_service.ensure_processing_time_column()
        self._browser_service = BrowserService(require_login_each_run=True)
        self._api_service = ApiService()

        self._settings_store = ReceiptSettingsStore(str(resolve_project_path(".runtime/receipt_settings.json")))
        self._receipt_settings: ReceiptSettings = self._settings_store.load()
        self._audio_service = WindowsAudioService()
        self._scan_success_sound_service = ScanSuccessSoundService(audio_service=self._audio_service)
        self._ticket_debug_tools_service = TicketDebugToolsService()

        self._order_viewmodel = OrderViewModel(self._excel_service, self._browser_service)

        self._scanner_view = ScannerView(
            camera_index=camera_index,
            focus_mode=self._receipt_settings.scanner_focus_mode,
            manual_focus_value=self._receipt_settings.scanner_manual_focus_value,
        )
        self._order_view = OrderView() if self._show_order_window else None
        if self._order_view is not None:
            self._order_view.set_print_request_callback(self._handle_manual_print)

        self._last_qr_url = ""
        self._last_qr_timestamp = 0.0
        self._qr_repeat_cooldown_sec = 2.0

        self._status_listener: StatusListener | None = None
        self._camera_listener: Callable[[str], None] | None = None
        self._camera_status_listener: Callable[[str | None], None] | None = None
        self._order_listener: Callable[[Order], None] | None = None
        self._stop_requested = False
        self._relogin_requested = False
        self._control_lock = threading.Lock()

    def set_camera_frame_listener(self, listener: Callable[[str], None] | None) -> None:
        """외부 뷰에서 카메라 프레임을 수신한다."""
        self._camera_listener = listener
        self._scanner_view.on_frame_ready = listener

    def set_camera_status_listener(self, listener: Callable[[str | None], None] | None) -> None:
        """외부 뷰에서 카메라 복구 상태 메시지를 수신한다."""
        self._camera_status_listener = listener
        set_listener = getattr(self._scanner_view, "set_camera_status_listener", None)
        if callable(set_listener):
            set_listener(listener)

    def set_order_listener(self, listener: Callable[[Order], None] | None) -> None:
        """QR 스캔 시 주문 정보를 외부 대시보드로 전달한다."""
        self._order_listener = listener

    def set_status_listener(self, listener: StatusListener | None) -> None:
        """외부 대시보드에서 상태 이벤트를 구독한다."""
        self._status_listener = listener

    def change_camera(self, new_index: int) -> None:
        """런타임 중 카메라 디바이스를 변경한다."""
        self._scanner_view.change_camera(new_index)

    def open_witchform_login_page(self) -> bool:
        """현재 주문 페이지와 별도로 설정된 로그인 페이지를 연다."""
        return bool(
            self._browser_service.open_page(
                self._browser_service.login_url,
                preserve_current_page=True,
            )
        )

    def apply_scanner_focus_settings(self, focus_mode: str, manual_focus_value: float | None) -> str:
        """실행 중인 스캐너에 초점 설정을 즉시 반영한다."""
        requested_manual = focus_mode == "manual"
        try:
            normalized_value = None if manual_focus_value is None else float(manual_focus_value)
        except (TypeError, ValueError):
            normalized_value = None
        invalid_manual_value = (
            requested_manual
            and normalized_value is not None
            and not math.isfinite(normalized_value)
        )
        if invalid_manual_value:
            normalized_value = None
        normalized_mode = "manual" if requested_manual and normalized_value is not None else "auto"

        scanner = getattr(self, "_scanner_view", None)
        capability_getter = getattr(scanner, "get_focus_capability", None)
        capability = capability_getter() if callable(capability_getter) else None
        manual_supported = None if capability is None else getattr(capability, "manual_focus_supported", None)
        if requested_manual and manual_supported is False:
            normalized_mode = "auto"
            normalized_value = None

        self._receipt_settings.scanner_focus_mode = normalized_mode
        self._receipt_settings.scanner_manual_focus_value = normalized_value

        if scanner is None:
            return "카메라 초점 설정 저장 완료 (다음 앱 시작 후 적용)"

        is_camera_ready = getattr(scanner, "is_camera_ready", None)
        camera_ready = bool(is_camera_ready()) if callable(is_camera_ready) else False

        apply_focus_settings = getattr(scanner, "apply_focus_settings", None)
        apply_result = (
            apply_focus_settings(normalized_mode, normalized_value)
            if callable(apply_focus_settings)
            else False
        )
        applied = bool(apply_result)
        apply_status = getattr(
            apply_result,
            "status",
            "applied" if applied else "failed",
        )

        updated_capability = capability_getter() if callable(capability_getter) else capability
        updated_manual_supported = (
            None
            if updated_capability is None
            else getattr(updated_capability, "manual_focus_supported", None)
        )
        if requested_manual and normalized_mode == "manual" and apply_status == "unsupported":
            self._receipt_settings.scanner_focus_mode = "auto"
            self._receipt_settings.scanner_manual_focus_value = None

        if requested_manual and manual_supported is False:
            return "현재 카메라가 수동 초점을 지원하지 않아 자동 초점으로 유지됩니다."
        if invalid_manual_value:
            return "수동 초점 값이 올바르지 않아 자동 초점으로 유지됩니다."
        if requested_manual and normalized_value is None:
            return "수동 초점 값이 없어 자동 초점으로 유지됩니다."
        if requested_manual and updated_manual_supported is False:
            return "현재 카메라가 수동 초점을 지원하지 않아 자동 초점으로 유지됩니다."
        if requested_manual and apply_status == "failed":
            return "수동 초점을 적용하지 못했습니다. 다시 시도하거나 카메라 고급 설정을 확인하세요."
        if camera_ready and applied:
            return "카메라 초점 설정 저장 완료 (현재 런타임에 바로 적용됨)"
        if not camera_ready:
            return "카메라 초점 설정 저장 완료 (카메라 준비 후 현재 런타임에 적용됨)"
        return "카메라 초점 설정 저장 완료 (현재 카메라가 이 설정을 바로 적용하지 못해 저장만 완료됨)"

    def get_scanner_focus_capability(self):
        """실행 중인 스캐너의 초점 capability를 반환한다."""
        scanner = getattr(self, "_scanner_view", None)
        if scanner is None:
            return None
        getter = getattr(scanner, "get_focus_capability", None)
        if not callable(getter):
            return None
        return getter()

    def open_scanner_camera_settings(self) -> bool:
        """현재 스캔 카메라의 Windows/제조사 설정 창을 엽니다."""
        scanner = getattr(self, "_scanner_view", None)
        if scanner is None:
            return False
        opener = getattr(scanner, "open_camera_settings", None)
        if not callable(opener):
            return False
        return bool(opener())

    def _get_control_lock(self) -> threading.Lock:
        lock = getattr(self, "_control_lock", None)
        if lock is None:
            lock = threading.Lock()
            self._control_lock = lock
        return lock

    def request_stop(self) -> None:
        """외부에서 런타임 정지를 요청한다.

        플래그만 설정하고 실제 자원 정리는 run()의 finally에서 수행한다.
        중복 stop 호출로 인한 Playwright 이벤트 루프 오류를 방지한다.
        """
        with self._get_control_lock():
            self._stop_requested = True
            self._relogin_requested = False
        self._emit_status("STOPPING", "런타임 중지 요청됨")

    def request_relogin(self) -> None:
        """외부에서 재로그인을 요청한다."""
        with self._get_control_lock():
            if self._stop_requested:
                return
            self._relogin_requested = True
        self._emit_status("RECOVERING", "재로그인 요청됨")

    def run(self) -> None:
        with self._get_control_lock():
            self._stop_requested = False
            self._relogin_requested = False
        self._emit_status("STARTING", "티켓 확인 런타임 시작 중")

        try:
            offline_mode = self._load_ticket_debug_settings().offline_scan_mode
            # 카메라를 가장 먼저 시작해 브라우저/뷰 초기화와 병렬로 준비되도록 한다
            self._scanner_view.start()
            if not offline_mode:
                self._browser_service.start()
            if self._order_view is not None:
                self._order_view.start()

            # 카메라 비동기 초기화 대기 (최대 30초)
            if not self._wait_for_camera(timeout_sec=30):
                self._enter_error("카메라를 열 수 없습니다. 장치를 확인해주세요.")
                return

            if offline_mode:
                self._enter_ready("[오프라인] 로그인 없이 스캔 테스트 모드")
            else:
                self._enter_auth_wait("브라우저에서 로그인이 필요합니다")
                if not self._wait_for_initial_login():
                    print("RUN_EXIT reason=login_failed_or_stop_requested")
                    return

            self._main_loop()
            print("RUN_EXIT reason=main_loop_finished")
        finally:
            self._emit_status("STOPPING", "런타임 종료 중")
            try:
                self._scanner_view.release()
            except Exception:
                pass
            try:
                if self._order_view is not None:
                    self._order_view.stop()
            except Exception:
                pass
            try:
                self._browser_service.stop()
            except Exception:
                pass
            self._emit_status("STOPPED", "런타임 종료됨")

    def _wait_for_camera(self, timeout_sec: int = 30) -> bool:
        """카메라 비동기 초기화 완료를 대기한다."""
        deadline = time.monotonic() + timeout_sec
        while time.monotonic() < deadline:
            if self._is_stop_requested():
                return False
            if self._scanner_view.is_camera_ready():
                return True
            time.sleep(0.1)
        return self._scanner_view.is_camera_ready()

    def _is_stop_requested(self) -> bool:
        with self._get_control_lock():
            return bool(getattr(self, "_stop_requested", False))

    def _consume_relogin_requested(self) -> bool:
        with self._get_control_lock():
            if getattr(self, "_stop_requested", False):
                self._relogin_requested = False
                return False
            if not getattr(self, "_relogin_requested", False):
                return False
            self._relogin_requested = False
            return True

    def _emit_order(self, order: Order) -> None:
        """주문 정보를 외부 대시보드로 전달한다."""
        listener = getattr(self, "_order_listener", None)
        if listener is None:
            return
        try:
            listener(order)
        except Exception:
            logger.warning("주문 콜백 처리 실패", exc_info=True)

    def _emit_status(self, state: str, message: str) -> None:
        listener = getattr(self, "_status_listener", None)
        if listener is None:
            return
        try:
            listener(state, message)
        except Exception:
            logger.warning("상태 콜백 처리 실패", exc_info=True)

    def _enter_auth_wait(self, message: str = "브라우저에서 로그인이 필요합니다") -> None:
        self._state = AppState.AUTH_WAIT
        self._scanner_view.set_auth_ready(False)
        self._scanner_view.set_scanning_enabled(True)
        self._scanner_view.set_status_message(message)
        self._emit_status(self._state.name, message)

    def _enter_ready(self, message: str = "준비됨") -> None:
        self._state = AppState.READY
        self._scanner_view.set_auth_ready(True)
        self._scanner_view.set_scanning_enabled(True)
        self._scanner_view.set_status_message(message)
        self._emit_status(self._state.name, message)

    def _enter_processing(self, message: str = "처리 중...") -> None:
        self._state = AppState.PROCESSING
        self._scanner_view.set_scanning_enabled(False)
        self._scanner_view.set_status_message(message)
        self._emit_status(self._state.name, message)

    def _enter_recovering(self, message: str = "로그인이 필요합니다 - 브라우저에서 로그인 완료하세요") -> None:
        self._state = AppState.RECOVERING
        self._scanner_view.set_auth_ready(False)
        self._scanner_view.set_scanning_enabled(True)
        self._scanner_view.set_status_message(message)
        self._emit_status(self._state.name, message)

    def _enter_error(self, message: str, keep_auth_state: bool = True) -> None:
        self._state = AppState.ERROR
        if not keep_auth_state:
            self._scanner_view.set_auth_ready(False)
        self._scanner_view.set_scanning_enabled(True)
        self._scanner_view.set_status_message(message)
        self._emit_status(self._state.name, message)

    def _browser_status_message(self, browser_result: BrowserResolveResult) -> str:
        if browser_result.error_code == "NETWORK_TIMEOUT":
            return "브라우저 응답이 지연되고 있습니다. 다시 스캔해주세요."
        if browser_result.error_code == "AUTH_REQUIRED":
            return "로그인이 필요합니다. 브라우저에서 로그인해주세요."
        return "브라우저 요청을 확인하지 못했습니다. 다시 스캔해주세요."

    def _receipt_failure_status_message(self, click_result: ReceiptClickResult) -> str:
        if click_result.error_code == "ALREADY_RECEIVED":
            return "이미 수령완료된 주문입니다 (주문정보만 표시)"
        if click_result.error_code in {"PRIMARY_CLICK_FAIL", "CONFIRM_CLICK_FAIL"}:
            return "수령 완료 처리 상태를 확인하지 못했습니다. 다시 스캔해주세요."
        if click_result.error_code == "NO_ACTIVE_PAGE":
            return "주문 상세 페이지를 확인하지 못했습니다. 다시 스캔해주세요."
        return "수령 완료 처리에 실패했습니다. 다시 스캔해주세요."

    def _handle_manual_print(self, order: Order) -> None:
        """구매자 정보 창에서 수동 영수증 출력 요청을 처리한다."""
        try:
            self._receipt_settings = self._settings_store.load()
        except Exception:
            logger.warning("영수증 설정 재로딩 실패", exc_info=True)
        try:
            print_order_receipt(order, self._receipt_settings)
            self._emit_status("READY", "영수증 수동 출력 완료")
        except Exception as exc:
            self._emit_status("ERROR", f"영수증 수동 출력 실패: {exc}")

    def _play_scan_success_sound(
        self,
        order: Order | None = None,
        *,
        increment_count: bool = True,
        persist_count: bool = True,
    ):
        settings_store = getattr(self, "_settings_store", None)
        if settings_store is not None:
            try:
                self._receipt_settings = settings_store.load()
            except Exception:
                logger.warning("스캔 사운드 설정 재로딩 실패", exc_info=True)

        receipt_settings = getattr(self, "_receipt_settings", None)
        sound_service = getattr(self, "_scan_success_sound_service", None)
        if sound_service is None:
            audio_service = getattr(self, "_audio_service", None)
            if audio_service is None:
                return
            sound_service = ScanSuccessSoundService(audio_service=audio_service)
            self._scan_success_sound_service = sound_service

        try:
            return sound_service.play_for_scan_success(
                receipt_settings,
                order_number="" if order is None else order.order_number,
                increment_count=increment_count,
                persist_count=persist_count,
            )
        except Exception as exc:
            print(f"SOUND_PLAY_FAIL order={'' if order is None else order.order_number} error={exc}")
            return None

    def _commit_scan_success_count(self) -> None:
        sound_service = getattr(self, "_scan_success_sound_service", None)
        if sound_service is None:
            return
        try:
            current_count = sound_service.load_success_count()
            sound_service.save_success_count(current_count + 1)
        except Exception:
            logger.warning("스캔 성공 누적 카운트 저장 실패", exc_info=True)

    def _load_ticket_debug_settings(self):
        service = getattr(self, "_ticket_debug_tools_service", None)
        if service is None:
            service = TicketDebugToolsService()
            self._ticket_debug_tools_service = service
        try:
            return service.load_settings()
        except Exception:
            logger.warning("디버그 설정 로딩 실패", exc_info=True)
            return TicketDebugSettings()

    def _should_play_duplicate_received_sound(self, debug_settings: TicketDebugSettings | None = None) -> bool:
        service = getattr(self, "_ticket_debug_tools_service", None)
        if service is None:
            service = TicketDebugToolsService()
            self._ticket_debug_tools_service = service
        try:
            return service.should_play_sound_for_duplicate_received_qr(debug_settings)
        except Exception:
            logger.warning("디버그 중복 스캔 효과음 설정 확인 실패", exc_info=True)
            return False

    def _should_count_scan_success_as_processed(self, debug_settings: TicketDebugSettings | None = None) -> bool:
        service = getattr(self, "_ticket_debug_tools_service", None)
        if service is None:
            service = TicketDebugToolsService()
            self._ticket_debug_tools_service = service
        try:
            return service.should_count_scan_success_as_processed(debug_settings)
        except Exception:
            logger.warning("디버그 누적 카운트 반영 설정 확인 실패", exc_info=True)
            return False

    def _resolve_qr_offline(self, qr_url: str) -> BrowserResolveResult:
        """브라우저/인증 없이 QR URL 리다이렉트를 추적한다 (오프라인 테스트 전용).

        test_order 파라미터가 있으면 HTTP 요청 없이 즉시 가짜 리다이렉트를 반환한다.
        """
        from urllib.parse import parse_qs, urlparse as _urlparse
        _parsed = _urlparse(qr_url)
        test_order = parse_qs(_parsed.query).get("test_order", [""])[0].strip()
        if test_order:
            return BrowserResolveResult(
                ok=True,
                status_code=302,
                location=f"/w/myform/sellForm-history-detail/{test_order}",
            )
        try:
            import httpx
            resp = httpx.get(qr_url, follow_redirects=True, timeout=5.0)
            final_url = str(resp.url)
            # 최종 도착 URL이 로그인 페이지이면 AUTH_REQUIRED 처리 경로로 유도
            if "/w/login" in final_url:
                return BrowserResolveResult(ok=True, status_code=302, location=final_url)
            return BrowserResolveResult(ok=True, status_code=resp.status_code, location=final_url)
        except Exception as exc:
            return BrowserResolveResult(
                ok=False, error_code="OFFLINE_HTTP_FAIL", error_message=str(exc)
            )

    def _extract_order_from_login_redirect(self, location: str) -> QRParseResult | None:
        """비인증 리다이렉트(/w/login?redirect=...)에서 주문번호를 추출한다."""
        from urllib.parse import parse_qs, unquote, urlparse
        parsed = urlparse(location)
        redirect_raw = parse_qs(parsed.query).get("redirect", [""])[0]
        if not redirect_raw:
            return None
        redirect_url = unquote(redirect_raw)
        parsed_redirect = urlparse(redirect_url)
        order_number = ApiService._extract_order_number(
            parsed_redirect.path, parsed_redirect.query, parsed_redirect.fragment
        )
        if not order_number:
            return None
        full_url = (
            f"{ApiService.WITCHFORM_BASE}{redirect_url}"
            if redirect_url.startswith("/")
            else redirect_url
        )
        return QRParseResult(success=True, order_number=order_number, full_url=full_url)

    def _wait_for_initial_login(self) -> bool:
        return self._wait_for_login(timeout_sec=0)

    def _wait_for_login(self, timeout_sec: int) -> bool:
        deadline = None if timeout_sec <= 0 else (time.monotonic() + timeout_sec)

        while not self._is_stop_requested():
            if deadline is not None and time.monotonic() >= deadline:
                break

            try:
                is_authenticated = self._browser_service.wait_until_authenticated(timeout_sec=1)
            except Exception as exc:
                if self._is_stop_requested():
                    return False
                self._enter_error(f"로그인 확인 실패: {exc}")
                time.sleep(0.2)
                continue

            if is_authenticated:
                self._enter_ready()
                return True

        if not self._is_stop_requested():
            self._enter_auth_wait("로그인 대기 시간 초과")
        return False

    def _handle_relogin_request(self) -> None:
        if self._is_stop_requested():
            return

        self._enter_recovering("재로그인 요청됨 - 브라우저에서 로그인하세요")
        try:
            ok = self._browser_service.request_relogin()
        except Exception as exc:
            self._enter_error(f"재로그인 요청 실패: {exc}")
            return

        if not ok:
            self._enter_error("재로그인 요청 실패")
            return

        self._wait_for_login(timeout_sec=180)

    def _main_loop(self) -> None:
        while self._scanner_view.is_running() and not self._is_stop_requested():
            if self._consume_relogin_requested():
                self._handle_relogin_request()
                continue

            qr_url = self._scanner_view.get_next_qr(timeout_sec=0.1)
            if not qr_url:
                continue

            if self._is_duplicate_qr(qr_url):
                self._scanner_view.set_status_message("중복 QR 코드 무시됨")
                self._scanner_view.set_scanning_enabled(True)
                continue

            self._enter_processing()
            self._process_qr(qr_url, allow_auth_retry=True)

        # 루프 탈출 원인 기록
        if not self._scanner_view.is_running():
            print("MAIN_LOOP_EXIT reason=camera_stopped")
        if self._is_stop_requested():
            print("MAIN_LOOP_EXIT reason=stop_requested")

    def _is_duplicate_qr(self, qr_url: str) -> bool:
        now = time.monotonic()
        is_duplicate = (
            qr_url == self._last_qr_url
            and (now - self._last_qr_timestamp) < self._qr_repeat_cooldown_sec
        )
        self._last_qr_url = qr_url
        self._last_qr_timestamp = now
        return is_duplicate

    def _process_qr(self, qr_url: str, allow_auth_retry: bool) -> None:
        if self._is_stop_requested():
            return
        normalized_qr_url = (qr_url or "").strip()
        parsed_qr = urlparse(normalized_qr_url)
        if parsed_qr.scheme not in {"http", "https"} or not parsed_qr.netloc:
            self._enter_error("유효한 QR URL이 아닙니다.")
            return
        if not normalized_qr_url.lower().startswith(ApiService.QR_PREFIX.lower()):
            self._enter_error("유효한 Witchform QR 코드가 아닙니다.")
            return
        offline_mode = self._load_ticket_debug_settings().offline_scan_mode
        if offline_mode:
            browser_result = self._resolve_qr_offline(normalized_qr_url)
        else:
            try:
                browser_result = self._browser_service.resolve_qr_redirect(normalized_qr_url)
            except Exception:
                if not self._is_stop_requested():
                    logger.warning("QR 브라우저 요청 실패", exc_info=True)
                    self._enter_error("브라우저 요청을 확인하지 못했습니다. 다시 스캔해주세요.")
                return

        self._process_resolved_qr(normalized_qr_url, browser_result, allow_auth_retry)

    def _process_resolved_qr(
        self,
        qr_url: str,
        browser_result: BrowserResolveResult,
        allow_auth_retry: bool,
    ) -> None:
        if self._is_stop_requested():
            return

        if not browser_result.ok:
            self._handle_browser_failure(browser_result, qr_url, allow_auth_retry)
            return

        parse_result = self._api_service.parse_qr_redirect(
            qr_url,
            browser_result.status_code,
            browser_result.location,
        )

        offline_mode = self._load_ticket_debug_settings().offline_scan_mode

        if not parse_result.success:
            print(
                "QR_PARSE_FAIL "
                f"code={parse_result.error_code} "
                f"status={browser_result.status_code} "
                f"location={browser_result.location}"
            )
            if parse_result.error_code == "AUTH_REQUIRED" and offline_mode:
                # 비인증 리다이렉트(/w/login?redirect=...)에서 주문번호 추출 시도
                recovered = self._extract_order_from_login_redirect(browser_result.location)
                if recovered:
                    parse_result = recovered
                else:
                    self._enter_error("오프라인 모드: 주문번호를 QR URL에서 추출하지 못했습니다.")
                    return
            elif parse_result.error_code == "AUTH_REQUIRED" and allow_auth_retry:
                self._recover_auth_and_retry(qr_url)
                return
            elif parse_result.error_code == "ORDER_NUMBER_MISSING":
                recovered_parse_result = self._recover_missing_order_number(browser_result)
                if recovered_parse_result is not None:
                    parse_result = recovered_parse_result
                else:
                    self._enter_error(parse_result.error_message)
                    return
            else:
                self._enter_error(parse_result.error_message)
                return

        order = self._order_viewmodel.load_order(
            parse_result.order_number,
            parse_result.full_url,
            open_page=False,
        )

        if not order:
            data_path = ensure_managed_data_file()
            logger.warning("주문번호 조회 실패 order=%s file=%s", parse_result.order_number, data_path)
            self._enter_error(f"주문번호를 찾을 수 없습니다: {parse_result.order_number}")
            return

        if (order.order_status or "").strip() in {"주문취소", "자동주문취소"}:
            self._enter_ready("취소된 주문은 수령 처리할 수 없습니다.")
            return

        if self._order_view is not None:
            self._order_view.show_or_update(order)
        self._emit_order(order)

        if order.is_received:
            debug_settings = self._load_ticket_debug_settings()
            count_as_processed = self._should_count_scan_success_as_processed(debug_settings)
            play_duplicate_sound = self._should_play_duplicate_received_sound(debug_settings)
            if count_as_processed or play_duplicate_sound:
                self._play_scan_success_sound(
                    order,
                    increment_count=count_as_processed,
                    persist_count=not count_as_processed,
                )
            if count_as_processed:
                self._commit_scan_success_count()
            self._enter_ready("이미 수령완료된 주문입니다 (주문정보만 표시)")
            return

        if not offline_mode:
            try:
                order_page_opened = self._order_viewmodel.open_current_order_page()
            except Exception:
                logger.warning("주문 상세 페이지 열기 중 예외 발생", exc_info=True)
                order_page_opened = False
            if not order_page_opened:
                self._enter_error("주문 상세 페이지를 열 수 없습니다.")
                return

            try:
                click_result = self._order_viewmodel.complete_receipt()
            except Exception as exc:
                logger.warning("수령 완료 처리 중 예외 발생", exc_info=True)
                click_result = ReceiptClickResult(
                    success=False, error_code="RECEIPT_EXCEPTION", error_message=str(exc)
                )
            if not click_result.success and click_result.error_code in {"PRIMARY_CLICK_FAIL", "CONFIRM_CLICK_FAIL"}:
                # 페이지 로딩 지연으로 인한 일시적 실패 가능성 — 1.5초 대기 후 1회 자동 재시도
                self._enter_processing("수령 완료 처리 재시도 중...")
                time.sleep(1.5)
                try:
                    retry_page_opened = self._order_viewmodel.open_current_order_page()
                except Exception:
                    logger.warning("수령 완료 재시도 페이지 열기 중 예외 발생", exc_info=True)
                    retry_page_opened = False
                if retry_page_opened:
                    try:
                        click_result = self._order_viewmodel.complete_receipt()
                    except Exception as exc:
                        logger.warning("수령 완료 재시도 중 예외 발생", exc_info=True)
                        click_result = ReceiptClickResult(
                            success=False, error_code="RECEIPT_EXCEPTION", error_message=str(exc)
                        )

            if not click_result.success:
                # 웹 페이지에서 이미 수령완료 상태인 경우 엑셀에 반영하고 정상 처리
                if click_result.error_code == "ALREADY_RECEIVED":
                    received_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self._order_viewmodel.mark_current_order_received(received_at)
                    debug_settings = self._load_ticket_debug_settings()
                    count_as_processed = self._should_count_scan_success_as_processed(debug_settings)
                    play_duplicate_sound = self._should_play_duplicate_received_sound(debug_settings)
                    if count_as_processed or play_duplicate_sound:
                        self._play_scan_success_sound(
                            order,
                            increment_count=count_as_processed,
                            persist_count=not count_as_processed,
                        )
                    if count_as_processed:
                        self._commit_scan_success_count()
                    self._enter_ready("이미 수령완료된 주문입니다 (주문정보만 표시)")
                    return
                if click_result.error_code in {"PRIMARY_CLICK_FAIL", "CONFIRM_CLICK_FAIL"}:
                    self._play_scan_success_sound(order, increment_count=False)
                logger.warning(
                    "수령 완료 처리 실패 code=%s message=%s",
                    click_result.error_code,
                    click_result.error_message,
                )
                self._enter_error(self._receipt_failure_status_message(click_result))
                return

        previous_received = order.received_at
        received_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not self._order_viewmodel.mark_current_order_received(received_at):
            data_path = ensure_managed_data_file()
            logger.warning("수령확인 저장 실패 order=%s file=%s", order.order_number, data_path)
            self._enter_error(f"수령확인 저장 실패: 엑셀 파일 권한/잠금을 확인해주세요. (파일: {data_path})")
            return

        # 표시용 진행상태 fallback과 분리해 앱 소유 주문상태 원본을 보존함
        previous_status = self._read_raw_order_status(order)
        try:
            status_saved = self._excel_service.mark_order_status(order.order_number, "거래종료")
        except Exception as exc:
            logger.warning("주문상태 저장 실패 order=%s error=%s", order.order_number, exc)
            status_saved = False
        if status_saved is False:
            rollback_ok = self._order_viewmodel.rollback_current_order_received(previous_received)
            if rollback_ok:
                self._enter_error("주문상태 저장 실패로 수령확인을 원복했습니다.")
            else:
                self._enter_error(f"주문상태 저장 및 수령확인 원복 실패 (수동 조치 필요, 주문번호={order.order_number})")
            return

        self._play_scan_success_sound(order, increment_count=True, persist_count=False)

        qr_auto_print_enabled = True
        try:
            # 설정 화면에서 변경된 최신 여백/템플릿 반영
            try:
                self._receipt_settings = self._settings_store.load()
            except Exception:
                pass
            qr_auto_print_enabled = bool(getattr(self._receipt_settings, "qr_scan_auto_print_enabled", True))
            if qr_auto_print_enabled:
                print_order_receipt(order, self._receipt_settings)
        except Exception as exc:
            rollback_ok = self._order_viewmodel.rollback_current_order_received(previous_received)
            if previous_status is not None:
                try:
                    self._excel_service.mark_order_status(order.order_number, previous_status)
                except Exception:
                    pass
            if rollback_ok:
                self._enter_error(f"영수증 출력 실패로 수령확인을 원복했습니다: {exc}")
            else:
                self._enter_error(
                    "영수증 출력 실패 및 수령확인 원복 실패 "
                    f"(수동 조치 필요, 주문번호={order.order_number}): {exc}"
                )
            return

        try:
            processing_time_writer = getattr(self._excel_service, "mark_order_processing_time", None)
            processing_time_saved = offline_mode or not callable(processing_time_writer) or processing_time_writer(order.order_number, received_at)
        except Exception as exc:
            logger.warning("처리시간 저장 실패 order=%s error=%s", order.order_number, exc)
            processing_time_saved = False
        if not processing_time_saved:
            self._enter_error("처리시간 저장 실패: 엑셀 파일 권한/잠금을 확인해주세요.")
            return

        self._commit_scan_success_count()
        if qr_auto_print_enabled:
            self._enter_ready("수령 완료 및 영수증 출력 완료")
        else:
            self._enter_ready("수령 완료 (QR 영수증 자동 출력 꺼짐)")

    def _read_raw_order_status(self, order: Order) -> str | None:
        """앱 소유 주문상태 컬럼의 원본 값을 반환한다."""
        file_path = getattr(self._excel_service, "_file_path", None)
        if not file_path:
            return None

        workbook = None
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook.active
            headers = {
                str(cell.value or "").strip().replace(" ", ""): index
                for index, cell in enumerate(worksheet[1], start=1)
            }
            order_column = headers.get("주문번호")
            status_column = headers.get("주문상태")
            if not order_column or not status_column:
                return ""
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if str(row[order_column - 1] or "").strip().upper() == order.order_number.upper():
                    return str(row[status_column - 1] or "").strip()
        except Exception:
            logger.warning("주문상태 원본 조회 실패 order=%s", order.order_number, exc_info=True)
        finally:
            if workbook is not None:
                workbook.close()
        return None

    def _recover_missing_order_number(
        self,
        browser_result: BrowserResolveResult,
    ) -> QRParseResult | None:
        detail_url = self._resolve_witchform_url(browser_result.location)
        if not detail_url:
            return None

        try:
            discovery = self._discover_page_order_context(detail_url)
        except Exception as exc:
            print(f"ORDER_DISCOVERY_FAIL url={detail_url} error={exc}")
            return None

        order_number = discovery["order_number"]
        if not order_number:
            recovered_order = self._recover_order_from_customer_context(
                discovery["buyer_name"],
                discovery["buyer_phone"],
            )
            if recovered_order is None:
                print(f"ORDER_DISCOVERY_MISS url={detail_url}")
                return None
            order_number = recovered_order.order_number
            print(
                "ORDER_DISCOVERY_CONTACT_RECOVERED "
                f"order_number={order_number} "
                f"buyer_name={_mask_name(discovery['buyer_name'])} "
                f"buyer_phone={_mask_phone(discovery['buyer_phone'])}"
            )

        recovered_url = discovery["url"] or detail_url
        print(f"ORDER_DISCOVERY_RECOVERED order_number={order_number} url={recovered_url}")
        return QRParseResult(
            success=True,
            order_number=order_number,
            full_url=recovered_url,
        )

    def _discover_page_order_context(self, detail_url: str) -> dict[str, str]:
        browser_service = self._browser_service
        if hasattr(browser_service, "discover_order_context_from_page"):
            discovery = browser_service.discover_order_context_from_page(detail_url)
            return {
                "order_number": (getattr(discovery, "order_number", "") or "").strip(),
                "buyer_name": (getattr(discovery, "buyer_name", "") or "").strip(),
                "buyer_phone": (getattr(discovery, "buyer_phone", "") or "").strip(),
                "url": (getattr(discovery, "url", "") or detail_url).strip(),
            }

        order_number = ""
        if hasattr(browser_service, "discover_order_number_from_page"):
            order_number = browser_service.discover_order_number_from_page(detail_url)
        return {
            "order_number": (order_number or "").strip(),
            "buyer_name": "",
            "buyer_phone": "",
            "url": detail_url,
        }

    def _recover_order_from_customer_context(self, buyer_name: str, buyer_phone: str) -> Order | None:
        buyer_name = (buyer_name or "").strip()
        buyer_phone = (buyer_phone or "").strip()
        if not buyer_name and not buyer_phone:
            return None

        excel_service = getattr(self, "_excel_service", None)
        if excel_service is None or not hasattr(excel_service, "find_orders_by_customer"):
            return None

        matches = excel_service.find_orders_by_customer(name=buyer_name, phone=buyer_phone)
        if not matches:
            print(
                "ORDER_DISCOVERY_CONTACT_MISS "
                f"buyer_name={_mask_name(buyer_name)} "
                f"buyer_phone={_mask_phone(buyer_phone)}"
            )
            return None
        if len(matches) != 1:
            print(
                "ORDER_DISCOVERY_CONTACT_AMBIGUOUS "
                f"buyer_name={_mask_name(buyer_name)} "
                f"buyer_phone={_mask_phone(buyer_phone)} "
                f"count={len(matches)}"
            )
            return None
        return matches[0]

    def _resolve_witchform_url(self, location: str) -> str:
        location = (location or "").strip()
        if not location:
            return ""
        if location.startswith(("http://", "https://")):
            return location
        base = getattr(self._api_service, "WITCHFORM_BASE", "https://witchform.com").rstrip("/")
        if location.startswith("/"):
            return f"{base}{location}"
        return f"{base}/{location.lstrip('/')}"

    def _handle_browser_failure(
        self,
        browser_result: BrowserResolveResult,
        qr_url: str,
        allow_auth_retry: bool,
    ) -> None:
        if self._is_stop_requested():
            return

        if browser_result.error_code == "AUTH_REQUIRED" and allow_auth_retry:
            self._recover_auth_and_retry(qr_url)
            return

        if browser_result.error_code == "NETWORK_TIMEOUT":
            try:
                retry_result = self._browser_service.resolve_qr_redirect(qr_url)
            except Exception as exc:
                logger.warning("QR 재시도 실패", exc_info=True)
                self._enter_error("브라우저 요청을 다시 확인하지 못했습니다. 다시 스캔해주세요.")
                return

            if retry_result.ok:
                self._process_resolved_qr(qr_url, retry_result, allow_auth_retry=False)
                return

            self._enter_error("네트워크 시간 초과가 발생했습니다")
            return

        logger.warning(
            "브라우저 처리 실패 code=%s message=%s",
            browser_result.error_code,
            browser_result.error_message,
        )
        self._enter_error(self._browser_status_message(browser_result))

    def _recover_auth_and_retry(self, qr_url: str) -> None:
        if self._is_stop_requested():
            return

        print("AUTH_REQUIRED")
        self._enter_recovering("로그인이 필요합니다 - 브라우저에서 로그인 완료하세요")

        if not self._wait_for_login(timeout_sec=180):
            if not self._is_stop_requested():
                self._enter_ready("로그인 대기 시간 초과 - 준비 상태로 돌아갑니다.")
            return

        self._enter_processing("로그인 확인 완료 - 다시 처리 중")
        self._process_qr(qr_url, allow_auth_retry=False)


if __name__ == "__main__":
    from views.dashboard_flet_view import run_dashboard_app

    run_dashboard_app()
